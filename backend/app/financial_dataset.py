from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook

try:
    import xlrd
except Exception:  # pragma: no cover
    xlrd = None  # type: ignore[assignment]

FieldKind = Literal["text", "numeric", "date"]


@dataclass(frozen=True)
class FieldSpec:
    key: str
    label: str
    kind: FieldKind
    aliases: tuple[str, ...] = ()


@dataclass
class ParsedWorkbookRow:
    sheet_name: str
    row_number: int
    values: dict[str, Any]
    raw_payload: dict[str, Any]


@dataclass
class ParsedWorkbook:
    rows: list[ParsedWorkbookRow]
    parsed_sheets: list[str]
    matched_columns: list[str]


FIELD_SPECS: tuple[FieldSpec, ...] = (
    FieldSpec("ms_ps", "MS/PS", "text", aliases=("MS PS", "MSPS")),
    FieldSpec(
        "classification",
        "Clasification",
        "text",
        aliases=("Classification",),
    ),
    FieldSpec("entity_as_per_gr", "Entity As per GR", "text"),
    FieldSpec("region", "Region", "text"),
    FieldSpec(
        "resource_id",
        "Emp ID",
        "text",
        aliases=("Resource ID", "Emp ID", "Employee ID", "Employee Id"),
    ),
    FieldSpec("resource_name", "Resource Name", "text"),
    FieldSpec("deal_type", "Deal Type", "text"),
    FieldSpec("bill_rate", "Bill Rate", "numeric"),
    FieldSpec("start_date", "Start Date", "date"),
    FieldSpec("end_date", "End Date", "date"),
    FieldSpec("apr_bgt", "AprBGT", "numeric"),
    FieldSpec("may_bgt", "MayBGT", "numeric"),
    FieldSpec("jun_bgt", "JunBGT", "numeric"),
    FieldSpec("jul_bgt", "JulBGT", "numeric"),
    FieldSpec("aug_bgt", "AugBGT", "numeric"),
    FieldSpec("sep_bgt", "SepBGT", "numeric"),
    FieldSpec("oct_bgt", "OctBGT", "numeric"),
    FieldSpec("nov_bgt", "NovBGT", "numeric"),
    FieldSpec("dec_bgt", "DecBGT", "numeric"),
    FieldSpec("jan_bgt", "JanBGT", "numeric"),
    FieldSpec("feb_bgt", "FebBGT", "numeric"),
    FieldSpec("mar_bgt", "MarBGT", "numeric"),
    FieldSpec("q1_bgt", "Q1BGT", "numeric"),
    FieldSpec("q2_bgt", "Q2BGT", "numeric"),
    FieldSpec("q3_bgt", "Q3BGT", "numeric"),
    FieldSpec("q4_bgt", "Q4BGT", "numeric"),
    FieldSpec("h1_bgt", "H1BGT", "numeric"),
    FieldSpec("h2_bgt", "H2BGT", "numeric"),
    FieldSpec("fy_bgt", "FYBGT", "numeric"),
    FieldSpec(
        "customer_name",
        "Customer Name",
        "text",
        aliases=("Updated Customer", "Updated Customer Name"),
    ),
    FieldSpec("project_name", "Project Name", "text"),
    FieldSpec("practice_head", "Practice Head", "text"),
    FieldSpec("bdm", "BDM", "text"),
    FieldSpec("remarks", "Remarks", "text"),
    FieldSpec("po_end_date", "PO End Date", "date"),
    FieldSpec("bu_head", "BU Head", "text"),
    FieldSpec("geo_head", "Geo Head", "text"),
    FieldSpec("business_type", "Business Type", "text"),
    FieldSpec("apr_fct", "Apr_FCT", "numeric"),
    FieldSpec("apr_act", "Apr_Act", "numeric", aliases=("Apr_ACT",)),
    FieldSpec("apr_var", "Apr_Var", "numeric", aliases=("Apr_VAR",)),
    FieldSpec("may_fct", "May_FCT", "numeric"),
    FieldSpec("may_act", "May_Act", "numeric", aliases=("May_ACT",)),
    FieldSpec("may_var", "May_Var", "numeric", aliases=("May_VAR",)),
    FieldSpec("jun_fct", "Jun_FCT", "numeric"),
    FieldSpec("jun_act", "Jun_Act", "numeric", aliases=("Jun_ACT",)),
    FieldSpec("jun_var", "Jun_Var", "numeric", aliases=("Jun_VAR",)),
    FieldSpec("q1_2025_26", "Q1 2025-26", "numeric"),
    FieldSpec("jul_fct", "Jul_FCT", "numeric"),
    FieldSpec("jul_act", "Jul_Act", "numeric", aliases=("Jul_ACT",)),
    FieldSpec("jul_var", "Jul_Var", "numeric", aliases=("Jul_VAR",)),
    FieldSpec("aug_fct", "Aug_FCT", "numeric"),
    FieldSpec("aug_act", "Aug_ACT", "numeric", aliases=("Aug_Act",)),
    FieldSpec("aug_var", "Aug_Var", "numeric", aliases=("Aug_VAR",)),
    FieldSpec("sep_fct", "Sep_FCT", "numeric"),
    FieldSpec("sep_act", "Sep_ACT", "numeric", aliases=("Sep_Act",)),
    FieldSpec("sep_var", "Sep_VAR", "numeric", aliases=("Sep_Var",)),
    FieldSpec("q2_2025_26", "Q2 2025-26", "numeric"),
    FieldSpec("h1", "H1", "numeric"),
    FieldSpec("oct_fct", "Oct_FCT", "numeric"),
    FieldSpec("oct_act", "Oct_ACT", "numeric", aliases=("Oct_Act",)),
    FieldSpec("oct_var", "Oct_Var", "numeric", aliases=("Oct_VAR",)),
    FieldSpec("nov_fct", "Nov_FCT", "numeric"),
    FieldSpec("nov_act", "Nov_ACT", "numeric", aliases=("Nov_Act",)),
    FieldSpec("nov_var", "Nov_Var", "numeric", aliases=("Nov_VAR",)),
    FieldSpec("dec_fct", "Dec_FCT", "numeric"),
    FieldSpec("dec_act", "Dec_Act", "numeric", aliases=("Dec_ACT",)),
    FieldSpec("dec_var", "Dec_Var", "numeric", aliases=("Dec_VAR",)),
    FieldSpec("q3_2025_26", "Q3 2025-26", "numeric"),
    FieldSpec("jan_fct", "Jan_FCT", "numeric"),
    FieldSpec("jan_act", "Jan_ACT", "numeric", aliases=("Jan_Act",)),
    FieldSpec("jan_var", "Jan_VAR", "numeric", aliases=("Jan_Var",)),
    FieldSpec("feb_fct", "Feb_FCT", "numeric"),
    FieldSpec("feb_act", "Feb_ACT", "numeric", aliases=("Feb_Act",)),
    FieldSpec("feb_var", "Feb_Var", "numeric", aliases=("Feb_VAR",)),
    FieldSpec("mar_fct", "Mar_FCT", "numeric"),
    FieldSpec("q4_2025_26", "Q4 2025-26", "numeric"),
    FieldSpec("h2", "H2", "numeric"),
    FieldSpec("fy_25_26", "FY-25-26", "numeric", aliases=("FY 25-26", "FY25-26")),
)

FIELD_BY_KEY = {field.key: field for field in FIELD_SPECS}
TEXT_FIELD_KEYS = tuple(field.key for field in FIELD_SPECS if field.kind == "text")
NUMERIC_FIELD_KEYS = tuple(field.key for field in FIELD_SPECS if field.kind == "numeric")
DATE_FIELD_KEYS = tuple(field.key for field in FIELD_SPECS if field.kind == "date")
REVENUE_COLUMN_KEYS = tuple(field.key for field in FIELD_SPECS)

MONTH_SEQUENCE = (
    ("Apr", "apr_bgt", "apr_fct", "apr_act", "apr_var"),
    ("May", "may_bgt", "may_fct", "may_act", "may_var"),
    ("Jun", "jun_bgt", "jun_fct", "jun_act", "jun_var"),
    ("Jul", "jul_bgt", "jul_fct", "jul_act", "jul_var"),
    ("Aug", "aug_bgt", "aug_fct", "aug_act", "aug_var"),
    ("Sep", "sep_bgt", "sep_fct", "sep_act", "sep_var"),
    ("Oct", "oct_bgt", "oct_fct", "oct_act", "oct_var"),
    ("Nov", "nov_bgt", "nov_fct", "nov_act", "nov_var"),
    ("Dec", "dec_bgt", "dec_fct", "dec_act", "dec_var"),
    ("Jan", "jan_bgt", "jan_fct", "jan_act", "jan_var"),
    ("Feb", "feb_bgt", "feb_fct", "feb_act", "feb_var"),
    ("Mar", "mar_bgt", "mar_fct", None, None),
)

SUMMARY_IDENTIFIERS = (
    "resource_id",
    "resource_name",
    "customer_name",
    "project_name",
    "region",
    "deal_type",
    "practice_head",
    "geo_head",
)

HEADER_MATCH_THRESHOLD = 8
DATE_FORMATS = (
    "%Y-%m-%d",
    "%d-%m-%Y",
    "%d/%m/%Y",
    "%m/%d/%Y",
    "%d-%b-%Y",
    "%d %b %Y",
    "%d.%m.%Y",
)


def normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.strip().lower())


KNOWN_HEADER_MAP: dict[str, str] = {}
for field in FIELD_SPECS:
    candidates = {field.label, field.key, *field.aliases}
    for candidate in candidates:
        normalized = normalize_header(candidate)
        if normalized:
            KNOWN_HEADER_MAP[normalized] = field.key


def parse_financial_workbook(path: Path) -> ParsedWorkbook:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _parse_openxml_workbook(path)
    if suffix == ".xls":
        return _parse_legacy_xls_workbook(path)
    raise ValueError("Upload an Excel workbook in .xlsx, .xlsm, or .xls format.")


def _parse_openxml_workbook(path: Path) -> ParsedWorkbook:
    workbook = load_workbook(filename=path, data_only=True, read_only=True)
    rows: list[ParsedWorkbookRow] = []
    parsed_sheets: list[str] = []
    matched_columns: set[str] = set()

    try:
        for worksheet in workbook.worksheets:
            raw_rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            sheet_rows, sheet_matches = _extract_sheet_rows(raw_rows, worksheet.title)
            if not sheet_rows:
                continue
            rows.extend(sheet_rows)
            parsed_sheets.append(worksheet.title)
            matched_columns.update(sheet_matches)
    finally:
        workbook.close()

    return ParsedWorkbook(
        rows=rows,
        parsed_sheets=parsed_sheets,
        matched_columns=sorted(matched_columns),
    )


def _parse_legacy_xls_workbook(path: Path) -> ParsedWorkbook:
    if xlrd is None:
        raise ValueError("The xlrd dependency is required for .xls workbook uploads.")

    workbook = xlrd.open_workbook(str(path), on_demand=True)
    rows: list[ParsedWorkbookRow] = []
    parsed_sheets: list[str] = []
    matched_columns: set[str] = set()

    for sheet in workbook.sheets():
        raw_rows: list[list[Any]] = []
        for row_index in range(sheet.nrows):
            row_values: list[Any] = []
            for column_index in range(sheet.ncols):
                row_values.append(
                    _coerce_xls_cell_value(
                        workbook,
                        sheet.cell(row_index, column_index),
                    )
                )
            raw_rows.append(row_values)

        sheet_rows, sheet_matches = _extract_sheet_rows(raw_rows, sheet.name)
        if not sheet_rows:
            continue
        rows.extend(sheet_rows)
        parsed_sheets.append(sheet.name)
        matched_columns.update(sheet_matches)

    return ParsedWorkbook(
        rows=rows,
        parsed_sheets=parsed_sheets,
        matched_columns=sorted(matched_columns),
    )


def _coerce_xls_cell_value(workbook: Any, cell: Any) -> Any:
    if xlrd is None:
        return cell.value

    if cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate_as_datetime(cell.value, workbook.datemode)
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        numeric_value = float(cell.value)
        return int(numeric_value) if numeric_value.is_integer() else numeric_value
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK, xlrd.XL_CELL_ERROR}:
        return None
    return cell.value


def _extract_sheet_rows(
    raw_rows: list[list[Any]],
    sheet_name: str,
) -> tuple[list[ParsedWorkbookRow], set[str]]:
    header_index, recognized_columns = _detect_header_row(raw_rows)
    if header_index is None or len(recognized_columns) < HEADER_MATCH_THRESHOLD:
        return [], set()

    headers = raw_rows[header_index]
    output: list[ParsedWorkbookRow] = []
    matched_keys = {field.key for field in recognized_columns.values()}

    for row_offset, raw_row in enumerate(raw_rows[header_index + 1 :], start=header_index + 2):
        if _is_blank_row(raw_row):
            continue

        values: dict[str, Any] = {}
        raw_payload: dict[str, Any] = {}

        for column_index, header_value in enumerate(headers):
            if column_index >= len(raw_row):
                continue

            header_text = str(header_value or "").strip()
            if not header_text:
                continue

            raw_cell_value = raw_row[column_index]
            raw_payload[header_text] = _serialize_value(raw_cell_value)

            field = recognized_columns.get(column_index)
            if field is None:
                continue

            coerced_value = _coerce_field_value(field.kind, raw_cell_value)
            if coerced_value is not None:
                values[field.key] = coerced_value

        if not _is_meaningful_row(values):
            continue

        output.append(
            ParsedWorkbookRow(
                sheet_name=sheet_name,
                row_number=row_offset,
                values=values,
                raw_payload=raw_payload,
            )
        )

    return output, matched_keys


def _detect_header_row(raw_rows: list[list[Any]]) -> tuple[int | None, dict[int, FieldSpec]]:
    best_index: int | None = None
    best_columns: dict[int, FieldSpec] = {}

    for row_index, row in enumerate(raw_rows[:25]):
        recognized_columns: dict[int, FieldSpec] = {}
        for column_index, cell_value in enumerate(row):
            normalized = normalize_header(str(cell_value or ""))
            if not normalized or normalized not in KNOWN_HEADER_MAP:
                continue

            field_key = KNOWN_HEADER_MAP[normalized]
            if any(existing.key == field_key for existing in recognized_columns.values()):
                continue

            recognized_columns[column_index] = FIELD_BY_KEY[field_key]

        if len(recognized_columns) > len(best_columns):
            best_index = row_index
            best_columns = recognized_columns

    return best_index, best_columns


def _is_blank_row(row: list[Any]) -> bool:
    return all(_is_empty_value(value) for value in row)


def _is_empty_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _coerce_field_value(kind: FieldKind, value: Any) -> Any:
    if _is_empty_value(value):
        return None
    if kind == "text":
        return _coerce_text_value(value)
    if kind == "numeric":
        return _coerce_numeric_value(value)
    return _coerce_date_value(value)


def _coerce_text_value(value: Any) -> str | None:
    if _is_empty_value(value):
        return None
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    text = str(value).strip()
    return text or None


def _coerce_numeric_value(value: Any) -> Decimal | None:
    if _is_empty_value(value):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))

    text = str(value).strip()
    if not text:
        return None

    negative = text.startswith("(") and text.endswith(")")
    cleaned = text.replace(",", "").replace("$", "").replace("%", "").strip()
    cleaned = cleaned.strip("()")
    cleaned = re.sub(r"[^0-9.\-]+", "", cleaned)
    if not cleaned:
        return None

    if negative and not cleaned.startswith("-"):
        cleaned = f"-{cleaned}"

    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def _coerce_date_value(value: Any) -> date | None:
    if _is_empty_value(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    for pattern in DATE_FORMATS:
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue

    return None


def _serialize_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def _is_meaningful_row(values: dict[str, Any]) -> bool:
    for key in SUMMARY_IDENTIFIERS:
        if values.get(key):
            return True
    return False
