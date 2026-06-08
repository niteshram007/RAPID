from __future__ import annotations

import csv
import os
import re
from pathlib import Path
from typing import BinaryIO

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook

try:
    import xlrd
except Exception:  # pragma: no cover
    xlrd = None  # type: ignore[assignment]

ALLOWED_UPLOAD_EXTENSIONS = {".xlsx", ".xls", ".csv"}
DEFAULT_MAX_UPLOAD_SIZE_BYTES = 1024 * 1024 * 1024
INVALID_FILENAME_CHARS_PATTERN = re.compile(r'[<>:"/\\|?*]')
WINDOWS_RESERVED_NAMES = {
    "con",
    "prn",
    "aux",
    "nul",
    "com1",
    "com2",
    "com3",
    "com4",
    "com5",
    "com6",
    "com7",
    "com8",
    "com9",
    "lpt1",
    "lpt2",
    "lpt3",
    "lpt4",
    "lpt5",
    "lpt6",
    "lpt7",
    "lpt8",
    "lpt9",
}


def get_max_upload_size_bytes() -> int:
    raw_value = os.getenv("RAPID_MAX_UPLOAD_BYTES", "").strip()
    if not raw_value:
        return DEFAULT_MAX_UPLOAD_SIZE_BYTES
    try:
        parsed = int(raw_value)
    except ValueError:
        return DEFAULT_MAX_UPLOAD_SIZE_BYTES
    return parsed if parsed > 0 else DEFAULT_MAX_UPLOAD_SIZE_BYTES


def validate_upload_filename(filename: str | None) -> str:
    raw_name = str(filename or "").strip()
    if not raw_name:
        raise HTTPException(status_code=400, detail="Select a workbook first.")
    if "/" in raw_name or "\\" in raw_name or ".." in raw_name:
        raise HTTPException(status_code=400, detail="Upload filename is not allowed.")
    if any(ord(character) < 32 for character in raw_name):
        raise HTTPException(status_code=400, detail="Upload filename is not allowed.")
    path = Path(raw_name)
    if path.name != raw_name or not path.stem.strip():
        raise HTTPException(status_code=400, detail="Upload filename is not allowed.")
    if path.stem.strip().lower() in WINDOWS_RESERVED_NAMES:
        raise HTTPException(status_code=400, detail="Upload filename is not allowed.")
    if len(path.suffixes) > 1:
        raise HTTPException(status_code=400, detail="Upload filename is not allowed.")
    if len(raw_name) > 180:
        raise HTTPException(status_code=400, detail="Upload filename is too long.")
    if INVALID_FILENAME_CHARS_PATTERN.search(raw_name):
        raise HTTPException(
            status_code=400,
            detail='Upload filename contains unsupported characters. Avoid < > : " / \\ | ? *.',
        )
    return raw_name


def validate_upload_extension(filename: str | None, dataset_type: str) -> str:
    safe_name = validate_upload_filename(filename)
    extension = Path(safe_name).suffix.lower()
    if extension not in ALLOWED_UPLOAD_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Only .xlsx, .xls, or .csv files are accepted.")
    if extension == ".csv" and dataset_type not in {"budget", "global_revenue"}:
        raise HTTPException(status_code=400, detail="CSV uploads are supported for Budget and Actuals files only.")
    return extension


def _is_formula_like_text(value: object) -> bool:
    if not isinstance(value, str):
        return False
    text = value.lstrip()
    if not text:
        return False
    if text[0] in {"=", "+", "@"}:
        return True
    return text.startswith("-") and len(text) > 1 and not text[1].isdigit()


def reject_formula_injection(path: Path) -> None:
    suffix = path.suffix.lower()
    # Spreadsheet uploads frequently contain legitimate formulas.
    # Keep import permissive for .xlsx/.xls and enforce safety on export/render.
    if suffix in {".xlsx", ".xls"}:
        return
    if suffix == ".csv":
        _reject_csv_formula_like_text(path)


def _reject_openxml_formula_like_values(path: Path) -> None:
    workbook = load_workbook(filename=path, data_only=True, read_only=True)
    try:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                for value in row:
                    if _is_formula_like_text(value):
                        raise HTTPException(
                            status_code=400,
                            detail="Upload rejected because it contains formula-like text values.",
                        )
    finally:
        workbook.close()


def _reject_xls_formula_like_text(path: Path) -> None:
    if xlrd is None:
        raise HTTPException(status_code=400, detail="The xlrd dependency is required for .xls uploads.")
    workbook = xlrd.open_workbook(str(path), on_demand=True)
    for sheet in workbook.sheets():
        for row_index in range(sheet.nrows):
            for column_index in range(sheet.ncols):
                value = sheet.cell_value(row_index, column_index)
                if _is_formula_like_text(value):
                    raise HTTPException(
                        status_code=400,
                        detail="Upload rejected because it contains formulas or formula-like values.",
                    )


def _reject_csv_formula_like_text(path: Path) -> None:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        for row in reader:
            for value in row:
                if _is_formula_like_text(value):
                    raise HTTPException(
                        status_code=400,
                        detail="Upload rejected because it contains formulas or formula-like values.",
                    )


def copy_upload_with_limit(source: BinaryIO, destination: Path) -> int:
    max_size = get_max_upload_size_bytes()
    total = 0
    with destination.open("wb") as output:
        while True:
            chunk = source.read(1024 * 1024)
            if not chunk:
                break
            total += len(chunk)
            if total > max_size:
                output.close()
                destination.unlink(missing_ok=True)
                raise HTTPException(
                    status_code=413,
                    detail=f"Upload file is too large. Limit is {max_size // (1024 * 1024)} MB.",
                )
            output.write(chunk)
    return total


def copy_upload_file(source: UploadFile, destination: Path) -> int:
    if hasattr(source.file, "seek"):
        source.file.seek(0)
    size = copy_upload_with_limit(source.file, destination)
    if size <= 0:
        destination.unlink(missing_ok=True)
        raise HTTPException(status_code=400, detail="Upload file is empty.")
    return size


def safe_store_upload(source: UploadFile, destination: Path) -> int:
    destination.parent.mkdir(parents=True, exist_ok=True)
    try:
        return copy_upload_file(source, destination)
    except HTTPException:
        raise
    except Exception as error:
        destination.unlink(missing_ok=True)
        raise HTTPException(status_code=400, detail="Unable to read uploaded file.") from error
