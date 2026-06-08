from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook

try:
    import xlrd
except Exception:  # pragma: no cover
    xlrd = None  # type: ignore[assignment]

FieldKind = Literal["text", "numeric", "date"]


@dataclass(frozen=True)
class RapidRevenueFieldSpec:
    key: str
    label: str
    kind: FieldKind
    aliases: tuple[str, ...] = ()


@dataclass
class RapidRevenueParsedRow:
    sheet_name: str
    row_number: int
    values: dict[str, Any]
    raw_payload: dict[str, Any]


@dataclass
class RapidRevenueParsedWorkbook:
    rows: list[RapidRevenueParsedRow]
    parsed_sheets: list[str]
    matched_columns: list[str]


RAPID_REVENUE_FIELD_SPECS: tuple[RapidRevenueFieldSpec, ...] = (
    RapidRevenueFieldSpec(
        "customer_name",
        "Customer Name",
        "text",
        aliases=(
            "Customer",
            "Customer name",
        ),
    ),
    RapidRevenueFieldSpec(
        "updated_customer",
        "Updated Customer",
        "text",
        aliases=("Updated Customer Name",),
    ),
    RapidRevenueFieldSpec(
        "ms_ps",
        "MS/PS",
        "text",
        aliases=("MS PS", "PS/MS", "PS/MS budget", "MS/PS budget"),
    ),
    RapidRevenueFieldSpec("entity", "Entity", "text", aliases=("Company",)),
    RapidRevenueFieldSpec(
        "gr_entity",
        "GR Entity",
        "text",
        aliases=("Entity As per GR", "Group company", "Group Company"),
    ),
    RapidRevenueFieldSpec(
        "row_us",
        "ROW/US",
        "text",
        aliases=("Region", "Region 2", "Region summary", "Region Summary"),
    ),
    RapidRevenueFieldSpec("strategic_account", "Strategic Account", "text"),
    RapidRevenueFieldSpec(
        "resource_id",
        "Emp ID",
        "text",
        aliases=("Resource ID", "Emp ID", "Employee ID", "Employee Id"),
    ),
    RapidRevenueFieldSpec("resource_name", "Resource Name", "text"),
    RapidRevenueFieldSpec("deal_type", "Deal Type", "text"),
    RapidRevenueFieldSpec("eeennn", "EEENNN", "text", aliases=("EENNN",)),
    RapidRevenueFieldSpec("bill_rate", "Bill Rate", "numeric"),
    RapidRevenueFieldSpec("rate_type", "Rate Type", "text", aliases=("Rate Type", "RateType")),
    RapidRevenueFieldSpec(
        "billed_currency",
        "Billed currency",
        "text",
        aliases=("Billed Currency", "Currency"),
    ),
    RapidRevenueFieldSpec(
        "forex",
        "Forex",
        "numeric",
        aliases=("FX", "FX Rate", "Forex Rate", "Exchange Rate"),
    ),
    RapidRevenueFieldSpec(
        "type_of_projects",
        "Type of Projects",
        "text",
        aliases=("Type of Project", "Type of Projects"),
    ),
    RapidRevenueFieldSpec("start_date", "Start Date", "date"),
    RapidRevenueFieldSpec("end_date", "End Date", "date"),
    RapidRevenueFieldSpec("apr_2026", "Apr 2026", "numeric", aliases=("Apr-2026", "AprBGT")),
    RapidRevenueFieldSpec("may_2026", "May 2026", "numeric", aliases=("May-2026", "MayBGT")),
    RapidRevenueFieldSpec("jun_2026", "Jun 2026", "numeric", aliases=("Jun-2026", "JunBGT")),
    RapidRevenueFieldSpec("jul_2026", "Jul 2026", "numeric", aliases=("Jul-2026", "JulBGT")),
    RapidRevenueFieldSpec("aug_2026", "Aug 2026", "numeric", aliases=("Aug-2026", "AugBGT")),
    RapidRevenueFieldSpec("sep_2026", "Sep 2026", "numeric", aliases=("Sep-2026", "SepBGT")),
    RapidRevenueFieldSpec("oct_2026", "Oct 2026", "numeric", aliases=("Oct-2026", "OctBGT")),
    RapidRevenueFieldSpec("nov_2026", "Nov 2026", "numeric", aliases=("Nov-2026", "NovBGT")),
    RapidRevenueFieldSpec("dec_2026", "Dec 2026", "numeric", aliases=("Dec-2026", "DecBGT")),
    RapidRevenueFieldSpec("jan_2027", "Jan 2027", "numeric", aliases=("Jan-2027", "JanBGT")),
    RapidRevenueFieldSpec("feb_2027", "Feb 2027", "numeric", aliases=("Feb-2027", "FebBGT")),
    RapidRevenueFieldSpec("mar_2027", "Mar 2027", "numeric", aliases=("Mar-2027", "MarBGT")),
    RapidRevenueFieldSpec("fy", "FY", "numeric", aliases=("FYBGT", "FY-25-26", "FY 25-26")),
    RapidRevenueFieldSpec("project_name", "Project Name", "text", aliases=("Project",)),
    RapidRevenueFieldSpec("client_name", "Client Name", "text"),
    RapidRevenueFieldSpec(
        "ocn_number",
        "OCN Number",
        "text",
        aliases=("OCN No", "OCN"),
    ),
    RapidRevenueFieldSpec("practice_head", "Practice Head", "text"),
    RapidRevenueFieldSpec("bdm", "BDM", "text"),
    RapidRevenueFieldSpec("geo_head", "Geo Head", "text", aliases=("GeoHead", "Geo head")),
    RapidRevenueFieldSpec("vertical", "Vertical", "text"),
    RapidRevenueFieldSpec("horizontal", "Horizontal", "text"),
    RapidRevenueFieldSpec("q1", "Q1", "numeric", aliases=("Q1BGT",)),
    RapidRevenueFieldSpec("q2", "Q2", "numeric", aliases=("Q2BGT",)),
    RapidRevenueFieldSpec("q3", "Q3", "numeric", aliases=("Q3BGT",)),
    RapidRevenueFieldSpec("q4", "Q4", "numeric", aliases=("Q4BGT",)),
)

RAPID_REVENUE_FIELD_BY_KEY = {
    field.key: field for field in RAPID_REVENUE_FIELD_SPECS
}
RAPID_REVENUE_COLUMN_KEYS = tuple(field.key for field in RAPID_REVENUE_FIELD_SPECS)
RAPID_REVENUE_JSON_LABELS = {
    field.key: field.label for field in RAPID_REVENUE_FIELD_SPECS
}
RAPID_REVENUE_SLICER_FIELDS = (
    "Practice Head",
    "BDM",
    "Geo Head",
    "Vertical",
    "Horizontal",
    "MS/PS",
    "Strategic Account",
    "Deal Type",
)
RAPID_REVENUE_MONTH_FIELDS = (
    "Apr 2026",
    "May 2026",
    "Jun 2026",
    "Jul 2026",
    "Aug 2026",
    "Sep 2026",
    "Oct 2026",
    "Nov 2026",
    "Dec 2026",
    "Jan 2027",
    "Feb 2027",
    "Mar 2027",
)
RAPID_REVENUE_NUMERIC_FIELDS = tuple(
    field.label for field in RAPID_REVENUE_FIELD_SPECS if field.kind == "numeric"
)
RAPID_REVENUE_REQUIRED_HEADER_THRESHOLD = 12
RAPID_REVENUE_DATE_FORMATS = (
    "%Y-%m-%d",
    "%d-%m-%Y",
    "%d/%m/%Y",
    "%m/%d/%Y",
    "%d-%b-%Y",
    "%d %b %Y",
    "%d.%m.%Y",
)


def normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.strip().lower())


_RAPID_HEADER_LOOKUP: dict[str, str] = {}
for field in RAPID_REVENUE_FIELD_SPECS:
    for candidate in {field.label, field.key, *field.aliases}:
        normalized = normalize_header(candidate)
        if normalized:
            _RAPID_HEADER_LOOKUP[normalized] = field.key


def parse_rapid_revenue_workbook(path: Path) -> RapidRevenueParsedWorkbook:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _parse_openxml_workbook(path)
    if suffix == ".xls":
        return _parse_legacy_xls_workbook(path)
    raise ValueError("Upload an Excel workbook in .xlsx, .xlsm, or .xls format.")


def _parse_openxml_workbook(path: Path) -> RapidRevenueParsedWorkbook:
    workbook = load_workbook(filename=path, data_only=True, read_only=True)
    rows: list[RapidRevenueParsedRow] = []
    parsed_sheets: list[str] = []
    matched_columns: set[str] = set()

    try:
        for worksheet in workbook.worksheets:
            sheet_rows, sheet_matches = _extract_sheet_rows_from_iterable(
                worksheet.iter_rows(values_only=True),
                worksheet.title,
            )
            if not sheet_rows:
                continue
            rows.extend(sheet_rows)
            parsed_sheets.append(worksheet.title)
            matched_columns.update(sheet_matches)
    finally:
        workbook.close()

    return RapidRevenueParsedWorkbook(
        rows=rows,
        parsed_sheets=parsed_sheets,
        matched_columns=sorted(matched_columns),
    )


def _extract_sheet_rows_from_iterable(
    row_iterable: Any,
    sheet_name: str,
) -> tuple[list[RapidRevenueParsedRow], set[str]]:
    preview_rows: list[list[Any]] = []
    output: list[RapidRevenueParsedRow] = []
    matched_keys: set[str] = set()
    headers: list[Any] | None = None
    recognized_columns: dict[int, RapidRevenueFieldSpec] | None = None

    for row_number, raw_row_tuple in enumerate(row_iterable, start=1):
        raw_row = list(raw_row_tuple)

        if headers is None or recognized_columns is None:
            if len(preview_rows) < 25:
                preview_rows.append(raw_row)
            header_index, candidate_columns = _detect_header_row(preview_rows)
            has_candidate_header = (
                header_index is not None
                and len(candidate_columns) >= RAPID_REVENUE_REQUIRED_HEADER_THRESHOLD
            )
            should_finalize_header = (
                has_candidate_header
                and header_index is not None
                and len(preview_rows) > header_index + 1
            ) or len(preview_rows) >= 25

            if not should_finalize_header:
                continue

            if not has_candidate_header or header_index is None:
                return [], set()

            headers = preview_rows[header_index]
            recognized_columns = candidate_columns
            matched_keys = {field.key for field in recognized_columns.values()}

            for preview_offset, preview_row in enumerate(preview_rows[header_index + 1 :]):
                preview_row_number = header_index + 2 + preview_offset
                parsed_row = _parse_sheet_row(
                    headers=headers,
                    recognized_columns=recognized_columns,
                    raw_row=preview_row,
                    sheet_name=sheet_name,
                    row_number=preview_row_number,
                )
                if parsed_row is not None:
                    output.append(parsed_row)
            continue

        parsed_row = _parse_sheet_row(
            headers=headers,
            recognized_columns=recognized_columns,
            raw_row=raw_row,
            sheet_name=sheet_name,
            row_number=row_number,
        )
        if parsed_row is not None:
            output.append(parsed_row)

    if headers is None or recognized_columns is None:
        header_index, candidate_columns = _detect_header_row(preview_rows)
        if (
            header_index is None
            or len(candidate_columns) < RAPID_REVENUE_REQUIRED_HEADER_THRESHOLD
        ):
            return [], set()
        headers = preview_rows[header_index]
        recognized_columns = candidate_columns
        matched_keys = {field.key for field in recognized_columns.values()}

        for preview_offset, preview_row in enumerate(preview_rows[header_index + 1 :]):
            preview_row_number = header_index + 2 + preview_offset
            parsed_row = _parse_sheet_row(
                headers=headers,
                recognized_columns=recognized_columns,
                raw_row=preview_row,
                sheet_name=sheet_name,
                row_number=preview_row_number,
            )
            if parsed_row is not None:
                output.append(parsed_row)

    return output, matched_keys


def _parse_legacy_xls_workbook(path: Path) -> RapidRevenueParsedWorkbook:
    if xlrd is None:
        raise ValueError("The xlrd dependency is required for .xls workbook uploads.")

    workbook = xlrd.open_workbook(str(path), on_demand=True)
    rows: list[RapidRevenueParsedRow] = []
    parsed_sheets: list[str] = []
    matched_columns: set[str] = set()

    for sheet in workbook.sheets():
        raw_rows: list[list[Any]] = []
        for row_index in range(sheet.nrows):
            row_values: list[Any] = []
            for column_index in range(sheet.ncols):
                row_values.append(_coerce_xls_cell_value(workbook, sheet.cell(row_index, column_index)))
            raw_rows.append(row_values)

        sheet_rows, sheet_matches = _extract_sheet_rows(raw_rows, sheet.name)
        if not sheet_rows:
            continue
        rows.extend(sheet_rows)
        parsed_sheets.append(sheet.name)
        matched_columns.update(sheet_matches)

    return RapidRevenueParsedWorkbook(
        rows=rows,
        parsed_sheets=parsed_sheets,
        matched_columns=sorted(matched_columns),
    )


def _coerce_xls_cell_value(workbook: Any, cell: Any) -> Any:
    if xlrd is None:
        return cell.value

    if cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate_as_datetime(cell.value, workbook.datemode)
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        numeric_value = float(cell.value)
        return int(numeric_value) if numeric_value.is_integer() else numeric_value
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK, xlrd.XL_CELL_ERROR}:
        return None
    return cell.value


def _extract_sheet_rows(
    raw_rows: list[list[Any]],
    sheet_name: str,
) -> tuple[list[RapidRevenueParsedRow], set[str]]:
    header_index, recognized_columns = _detect_header_row(raw_rows)
    if header_index is None or len(recognized_columns) < RAPID_REVENUE_REQUIRED_HEADER_THRESHOLD:
        return [], set()

    headers = raw_rows[header_index]
    output: list[RapidRevenueParsedRow] = []
    matched_keys = {field.key for field in recognized_columns.values()}

    for row_offset, raw_row in enumerate(raw_rows[header_index + 1 :], start=header_index + 2):
        if _is_blank_row(raw_row):
            continue

        values: dict[str, Any] = {}
        raw_payload: dict[str, Any] = {}

        for column_index, header_value in enumerate(headers):
            if column_index >= len(raw_row):
                continue

            header_text = str(header_value or "").strip()
            if not header_text:
                continue

            raw_cell_value = raw_row[column_index]
            raw_payload[header_text] = _serialize_value(raw_cell_value)

            field = recognized_columns.get(column_index)
            if field is None:
                continue

            coerced_value = _coerce_field_value(field.kind, raw_cell_value)
            if coerced_value is not None:
                values[field.key] = coerced_value

        if not _is_meaningful_row(values):
            continue

        output.append(
            RapidRevenueParsedRow(
                sheet_name=sheet_name,
                row_number=row_offset,
                values=values,
                raw_payload=raw_payload,
            )
        )

    return output, matched_keys


def _parse_sheet_row(
    *,
    headers: list[Any],
    recognized_columns: dict[int, RapidRevenueFieldSpec],
    raw_row: list[Any],
    sheet_name: str,
    row_number: int,
) -> RapidRevenueParsedRow | None:
    if _is_blank_row(raw_row):
        return None

    values: dict[str, Any] = {}
    raw_payload: dict[str, Any] = {}

    for column_index, header_value in enumerate(headers):
        if column_index >= len(raw_row):
            continue

        header_text = str(header_value or "").strip()
        if not header_text:
            continue

        raw_cell_value = raw_row[column_index]
        raw_payload[header_text] = _serialize_value(raw_cell_value)

        field = recognized_columns.get(column_index)
        if field is None:
            continue

        coerced_value = _coerce_field_value(field.kind, raw_cell_value)
        if coerced_value is not None:
            values[field.key] = coerced_value

    if not _is_meaningful_row(values):
        return None

    return RapidRevenueParsedRow(
        sheet_name=sheet_name,
        row_number=row_number,
        values=values,
        raw_payload=raw_payload,
    )


def _detect_header_row(
    raw_rows: list[list[Any]],
) -> tuple[int | None, dict[int, RapidRevenueFieldSpec]]:
    best_index: int | None = None
    best_columns: dict[int, RapidRevenueFieldSpec] = {}

    for row_index, row in enumerate(raw_rows[:25]):
        recognized_columns: dict[int, RapidRevenueFieldSpec] = {}
        for column_index, cell_value in enumerate(row):
            normalized = normalize_header(str(cell_value or ""))
            if not normalized or normalized not in _RAPID_HEADER_LOOKUP:
                continue

            field_key = _RAPID_HEADER_LOOKUP[normalized]
            if any(existing.key == field_key for existing in recognized_columns.values()):
                continue

            recognized_columns[column_index] = RAPID_REVENUE_FIELD_BY_KEY[field_key]

        if len(recognized_columns) > len(best_columns):
            best_index = row_index
            best_columns = recognized_columns

    return best_index, best_columns


def _is_blank_row(row: list[Any]) -> bool:
    return all(_is_empty_value(value) for value in row)


def _is_empty_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _coerce_field_value(kind: FieldKind, value: Any) -> Any:
    if _is_empty_value(value):
        return None
    if kind == "text":
        return _coerce_text_value(value)
    if kind == "numeric":
        return _coerce_numeric_value(value)
    return _coerce_date_value(value)


def _coerce_text_value(value: Any) -> str | None:
    if _is_empty_value(value):
        return None
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    text = str(value).strip()
    return text or None


def _coerce_numeric_value(value: Any) -> Decimal | None:
    if _is_empty_value(value) or isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))

    text = str(value).strip()
    if not text:
        return None

    negative = text.startswith("(") and text.endswith(")")
    cleaned = text.replace(",", "").replace("$", "").replace("%", "").strip()
    cleaned = cleaned.strip("()")
    cleaned = re.sub(r"[^0-9.\-]+", "", cleaned)
    if not cleaned:
        return None

    if negative and not cleaned.startswith("-"):
        cleaned = f"-{cleaned}"

    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def _coerce_date_value(value: Any) -> date | None:
    if _is_empty_value(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    for fmt in RAPID_REVENUE_DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _is_meaningful_row(values: dict[str, Any]) -> bool:
    # Keep any non-empty parsed row so uploads are not dropped when specific
    # identifiers (for example OCN or Resource ID) are blank.
    return any(value not in (None, "") for value in values.values())


def _serialize_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value

