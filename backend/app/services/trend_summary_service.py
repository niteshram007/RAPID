from __future__ import annotations

import io
import re
from datetime import date, datetime
from decimal import Decimal
from typing import Any

import pandas as pd
from openpyxl import Workbook

from ..postgres import ensure_postgres_schema, open_database_connection
from ..security import sanitize_export_cell
from .anomaly_detection_service import apply_anomaly_detection
from .budget_processing_service import refresh_budget_data
from .global_revenue_processing_service import refresh_actual_revenue
from .insight_generation_service import build_insights
from .ml_prediction_service import build_predictions
from .risk_scoring_service import apply_risk_scores
from .trend_common import QUARTER_BY_MONTH, month_sort_key, parse_financial_year, safe_pct, utc_now_iso

SUMMARY_TEXT_FIELDS = (
    "customer_name",
    "project_name",
    "bdm",
    "practice_head",
    "geo_head",
    "vertical",
    "horizontal",
    "entity",
    "ms_ps",
    "region",
    "sales_region",
    "strategic_account",
    "quarter",
    "prediction_method",
    "risk_reason",
)
SUMMARY_NUMERIC_FIELDS = (
    "budget_amount",
    "actual_revenue",
    "budget_variance",
    "budget_variance_percent",
    "budget_achievement_percent",
    "previous_month_revenue",
    "previous_year_revenue",
    "revenue_growth_percent",
    "year_over_year_growth_percent",
    "margin_amount",
    "margin_percent",
    "utilization_percent",
    "predicted_revenue",
    "prediction_confidence",
    "risk_score",
)
FILTER_COLUMN_MAP = {
    "financialYear": "fy_year",
    "financialYears": "fy_year",
    "fy_year": "fy_year",
    "month": "month",
    "quarter": "quarter",
    "year": "year",
    "customerName": "customer_name",
    "customerNames": "customer_name",
    "customer_name": "customer_name",
    "projectName": "project_name",
    "projectNames": "project_name",
    "project_name": "project_name",
    "bdms": "bdm",
    "bdm": "bdm",
    "practiceHead": "practice_head",
    "practiceHeads": "practice_head",
    "practice_head": "practice_head",
    "geoHead": "geo_head",
    "geoHeads": "geo_head",
    "geo_head": "geo_head",
    "verticals": "vertical",
    "vertical": "vertical",
    "horizontals": "horizontal",
    "horizontal": "horizontal",
    "msps": "ms_ps",
    "ms_ps": "ms_ps",
    "rowUs": "region",
    "regions": "region",
    "region": "region",
    "salesRegions": "sales_region",
    "sales_region": "sales_region",
    "entities": "entity",
    "entity": "entity",
    "strategicAccount": "strategic_account",
    "strategicAccounts": "strategic_account",
    "strategic_account": "strategic_account",
}
MONTH_ORDER = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
BUDGET_MONTH_FIELD_BY_NAME = {
    "Apr": "apr_2026",
    "May": "may_2026",
    "Jun": "jun_2026",
    "Jul": "jul_2026",
    "Aug": "aug_2026",
    "Sep": "sep_2026",
    "Oct": "oct_2026",
    "Nov": "nov_2026",
    "Dec": "dec_2026",
    "Jan": "jan_2027",
    "Feb": "feb_2027",
    "Mar": "mar_2027",
}
QUARTER_MONTHS = {
    "Q1": ("Apr", "May", "Jun"),
    "Q2": ("Jul", "Aug", "Sep"),
    "Q3": ("Oct", "Nov", "Dec"),
    "Q4": ("Jan", "Feb", "Mar"),
}
MIN_BUDGET_DENOMINATOR_FOR_PERCENT = 1.0
CURRENCY_ZERO_EPSILON = 0.005
NUMERIC_18_2_ABS_MAX = 9_999_999_999_999_999.0
US_REGION_ALIASES = {
    "US",
    "USA",
    "USN",
    "USW",
    "USE",
    "USS",
    "USC",
    "UNITED STATES",
    "UNITED STATES OF AMERICA",
}
ROW_REGION_ALIASES = {
    "ROW",
    "REST OF WORLD",
    "REST-OF-WORLD",
}


def refresh_trend_analytics(
    financial_year: str | None = None,
    upload_month: str | None = None,
) -> dict[str, Any]:
    ensure_postgres_schema()
    refreshed_years: list[str] = []
    total_budget_rows = 0
    total_actual_rows = 0
    total_summary_rows = 0
    total_predictions = 0
    total_insights = 0
    total_anomalies = 0

    with open_database_connection(require=True) as connection:
        assert connection is not None
        years = _resolve_financial_years(connection, financial_year)
        for active_year in years:
            budget_result = refresh_budget_data(active_year, connection)
            actual_result = refresh_actual_revenue(active_year, connection)
            summary_result = _refresh_trend_tables_for_year(active_year, connection)
            refreshed_years.append(active_year)
            total_budget_rows += int(budget_result.get("rows_processed") or 0)
            total_actual_rows += int(actual_result.get("rows_processed") or 0)
            total_summary_rows += int(summary_result.get("rows_processed") or 0)
            total_predictions += int(summary_result.get("predictions_generated") or 0)
            total_insights += int(summary_result.get("insights_generated") or 0)
            total_anomalies += int(summary_result.get("anomalies_detected") or 0)
        connection.commit()

    return {
        "status": "success",
        "financialYears": refreshed_years,
        "uploadMonth": str(upload_month or "").strip() or None,
        "budgetRowsProcessed": total_budget_rows,
        "actualRowsProcessed": total_actual_rows,
        "rowsProcessed": total_summary_rows,
        "summaryRefreshed": bool(refreshed_years),
        "riskScoresUpdated": bool(refreshed_years),
        "anomaliesDetected": total_anomalies,
        "predictionsGenerated": total_predictions,
        "insightsGenerated": total_insights,
        "refreshedAt": utc_now_iso(),
    }


def refresh_trend_summary(month: str | None = None, year: str | None = None) -> dict[str, Any]:
    _ = month
    return refresh_trend_analytics(financial_year=year)


def get_trend_filters(filters: dict[str, Any] | None = None) -> dict[str, Any]:
    ensure_postgres_schema()
    frame = _load_summary_frame(filters)
    if frame.empty:
        return {
            "financialYears": [],
            "months": MONTH_ORDER,
            "quarters": ["Q1", "Q2", "Q3", "Q4"],
            "customerNames": [],
            "projectNames": [],
            "bdms": [],
            "practiceHeads": [],
            "geoHeads": [],
            "verticals": [],
            "horizontals": [],
            "msps": [],
            "regions": [],
            "salesRegions": [],
            "strategicAccounts": [],
            "lastUpdated": None,
        }

    def distinct(column: str) -> list[str]:
        if column not in frame:
            return []
        values = [
            str(value).strip()
            for value in frame[column].fillna("").tolist()
            if str(value).strip()
        ]
        return sorted(dict.fromkeys(values), key=str.lower)

    return {
        "financialYears": distinct("fy_year"),
        "months": [month for month in MONTH_ORDER if month in set(distinct("month"))] or MONTH_ORDER,
        "quarters": [quarter for quarter in ["Q1", "Q2", "Q3", "Q4"] if quarter in set(distinct("quarter"))] or ["Q1", "Q2", "Q3", "Q4"],
        "customerNames": distinct("customer_name"),
        "projectNames": distinct("project_name"),
        "bdms": distinct("bdm"),
        "practiceHeads": distinct("practice_head"),
        "geoHeads": distinct("geo_head"),
        "verticals": distinct("vertical"),
        "horizontals": distinct("horizontal"),
        "msps": distinct("ms_ps"),
        "regions": distinct("region"),
        "salesRegions": distinct("sales_region"),
        "strategicAccounts": distinct("strategic_account"),
        "lastUpdated": _frame_last_updated(frame),
    }


def get_trend_kpis(filters: dict[str, Any] | None = None) -> dict[str, Any]:
    frame = _load_enriched_summary_frame(filters)
    if frame.empty:
        return {"cards": [], "lastUpdated": None}

    prediction_frame = _load_predictions_frame(filters, prediction_levels=["overall", "project"])
    previous_period = _previous_period_totals(frame)
    current_period = _latest_period_totals(frame)
    predicted_total = float(prediction_frame["predicted_revenue"].sum()) if not prediction_frame.empty else 0.0
    total_budget = float(frame["budget_amount"].sum())
    total_actual = float(frame["actual_revenue"].sum())
    total_variance = float(frame["budget_variance"].sum())
    achievement_pct = safe_pct(total_actual, total_budget)
    previous_total_actual = float(previous_period.get("actual_revenue") or 0.0)
    revenue_growth_pct = safe_pct(total_actual - previous_total_actual, previous_total_actual) if previous_total_actual > 0 else 0.0
    total_previous_year = float(frame["previous_year_revenue"].sum())
    yoy_growth_pct = safe_pct(total_actual - total_previous_year, total_previous_year) if total_previous_year > 0 else 0.0
    high_risk_count = int(frame.loc[frame["risk_level"] == "High"].shape[0])
    anomaly_count = int(frame["anomaly_flag"].fillna(False).astype(bool).sum())
    unplanned_revenue = float(
        frame.loc[(frame["budget_amount"] <= 0) & (frame["actual_revenue"] > 0), "actual_revenue"].sum()
    )

    cards = [
        _kpi_card(
            "Total Budget",
            total_budget,
            current_period.get("budget_amount"),
            previous_period.get("budget_amount"),
            "Budgeted revenue in the current filtered scope.",
            "currency",
        ),
        _kpi_card(
            "Total Actuals",
            total_actual,
            current_period.get("actual_revenue"),
            previous_period.get("actual_revenue"),
            "Realized revenue from the latest uploaded Actuals data.",
            "currency",
        ),
        _kpi_card(
            "Budget Achievement %",
            achievement_pct,
            current_period.get("budget_achievement_percent"),
            previous_period.get("budget_achievement_percent"),
            "Share of budget converted into actual revenue.",
            "percent",
        ),
        _kpi_card(
            "Budget Variance",
            total_variance,
            current_period.get("budget_variance"),
            previous_period.get("budget_variance"),
            "Positive means actuals are ahead of plan.",
            "currency",
        ),
        _kpi_card(
            "Month-over-Month Growth %",
            revenue_growth_pct,
            current_period.get("revenue_growth_percent"),
            previous_period.get("revenue_growth_percent"),
            "Change in actual revenue compared with the previous month.",
            "percent",
        ),
        _kpi_card(
            "Year-over-Year Growth %",
            yoy_growth_pct,
            current_period.get("year_over_year_growth_percent"),
            previous_period.get("year_over_year_growth_percent"),
            "Change in actual revenue compared with the same period in the prior year.",
            "percent",
        ),
        _kpi_card(
            "Predicted Revenue",
            predicted_total,
            predicted_total,
            current_period.get("actual_revenue"),
            "Future revenue summed from project-level prediction rows.",
            "currency",
        ),
        _kpi_card(
            "High Risk Count",
            float(high_risk_count),
            float(high_risk_count),
            None,
            "High risk rows in the current slice.",
            "count",
        ),
        _kpi_card(
            "Anomaly Count",
            float(anomaly_count),
            float(anomaly_count),
            None,
            "Detected revenue, budget, margin, or utilization anomalies.",
            "count",
        ),
        _kpi_card(
            "Unplanned Revenue",
            unplanned_revenue,
            unplanned_revenue,
            None,
            "Actual revenue delivered without a matching budget baseline.",
            "currency",
        ),
    ]

    return {
        "cards": cards,
        "lastUpdated": _frame_last_updated(frame),
    }


def get_monthly_comparison(filters: dict[str, Any] | None = None) -> dict[str, Any]:
    frame = _load_enriched_summary_frame(filters)
    if frame.empty:
        return {"rows": [], "lastUpdated": None}

    monthly = (
        frame.groupby(["month", "year", "quarter", "month_sort_key"], dropna=False)
        .agg({"budget_amount": "sum", "actual_revenue": "sum", "predicted_revenue": "sum"})
        .reset_index()
    )

    rows: list[dict[str, Any]] = []
    for _, row in monthly.sort_values("month_sort_key").iterrows():
        month = str(row["month"])
        year = int(row["year"])
        rows.append(
            {
                "month": month,
                "year": year,
                "quarter": str(row["quarter"]),
                "label": f"{month} {year}",
                "budget": float(row["budget_amount"]),
                "actual": float(row["actual_revenue"]),
                "predicted": float(row["predicted_revenue"]),
                "variance": float(row["actual_revenue"] - row["budget_amount"]),
                "variancePercent": safe_pct(
                    float(row["actual_revenue"] - row["budget_amount"]),
                    float(row["budget_amount"]),
                ),
            }
        )

    return {
        "rows": rows,
        "lastUpdated": _frame_last_updated(frame),
    }


def get_budget_vs_actual(filters: dict[str, Any] | None = None) -> dict[str, Any]:
    frame = _load_enriched_summary_frame(filters)
    if frame.empty:
        return {"breakdowns": {}, "lastUpdated": None}

    return {
        "breakdowns": {
            "bdm": _build_dimension_breakdown(frame, "bdm"),
            "practice_head": _build_dimension_breakdown(frame, "practice_head"),
            "geo_head": _build_dimension_breakdown(frame, "geo_head"),
            "vertical": _build_dimension_breakdown(frame, "vertical"),
            "horizontal": _build_dimension_breakdown(frame, "horizontal"),
            "ms_ps": _build_dimension_breakdown(frame, "ms_ps"),
            "customer": _build_dimension_breakdown(frame, "customer_name"),
        },
        "underperformers": _build_underperformer_rows(frame),
        "lastUpdated": _frame_last_updated(frame),
    }


def get_year_over_year(filters: dict[str, Any] | None = None) -> dict[str, Any]:
    frame = _load_enriched_summary_frame(filters)
    if frame.empty:
        return {"rows": [], "breakdowns": {}, "summary": None, "lastUpdated": None}

    monthly = (
        frame.groupby(["month", "year", "month_sort_key"], dropna=False)
        .agg({"actual_revenue": "sum", "previous_year_revenue": "sum"})
        .reset_index()
        .sort_values("month_sort_key")
    )
    rows = [
        {
            "month": str(row["month"]),
            "currentYear": int(row["year"]),
            "previousYear": int(row["year"]) - 1,
            "currentYearRevenue": float(row["actual_revenue"]),
            "previousYearRevenue": float(row["previous_year_revenue"]),
            "yoyVariance": float(row["actual_revenue"] - row["previous_year_revenue"]),
            "yoyVariancePercent": safe_pct(
                float(row["actual_revenue"] - row["previous_year_revenue"]),
                float(row["previous_year_revenue"]),
            )
            if float(row["previous_year_revenue"]) > 0
            else 0.0,
            "yoyGrowthStatus": _growth_status(
                float(row["actual_revenue"]),
                float(row["previous_year_revenue"]),
            ),
        }
        for _, row in monthly.iterrows()
    ]
    total_current = float(frame["actual_revenue"].sum())
    total_previous = float(frame["previous_year_revenue"].sum())
    return {
        "rows": rows,
        "breakdowns": {
            "customer": _build_yoy_breakdown(frame, "customer_name"),
            "bdm": _build_yoy_breakdown(frame, "bdm"),
            "practice_head": _build_yoy_breakdown(frame, "practice_head"),
            "geo_head": _build_yoy_breakdown(frame, "geo_head"),
            "ms_ps": _build_yoy_breakdown(frame, "ms_ps"),
        },
        "summary": {
            "currentYearRevenue": total_current,
            "previousYearRevenue": total_previous,
            "yoyVariance": total_current - total_previous,
            "yoyGrowthPercent": safe_pct(total_current - total_previous, total_previous)
            if total_previous > 0
            else 0.0,
        },
        "lastUpdated": _frame_last_updated(frame),
    }


def get_risk_data(filters: dict[str, Any] | None = None) -> dict[str, Any]:
    frame = _load_enriched_summary_frame(filters)
    if frame.empty:
        return {"rows": [], "lastUpdated": None}

    risk_frame = frame[
        [
            "customer_name",
            "project_name",
            "bdm",
            "practice_head",
            "geo_head",
            "vertical",
            "horizontal",
            "ms_ps",
            "month",
            "year",
            "budget_amount",
            "actual_revenue",
            "budget_variance_percent",
            "revenue_growth_percent",
            "year_over_year_growth_percent",
            "risk_score",
            "risk_level",
            "risk_reason",
        ]
    ].copy()
    risk_frame["reason"] = risk_frame["risk_reason"].fillna("").apply(_clean_text)
    risk_frame = risk_frame.sort_values(
        ["risk_score", "budget_variance_percent", "customer_name"],
        ascending=[False, True, True],
    )
    rows = [
        {
            "name": " / ".join(
                part
                for part in [str(row["customer_name"]).strip(), str(row["project_name"]).strip()]
                if part
            )
            or str(row["customer_name"]).strip()
            or "Unassigned",
            "type": "Project" if str(row["project_name"]).strip() else "Customer",
            "customerName": str(row["customer_name"]),
            "projectName": str(row["project_name"]),
            "month": str(row["month"]),
            "year": int(row["year"]),
            "budget": float(row["budget_amount"]),
            "actual": float(row["actual_revenue"]),
            "variancePercent": float(row["budget_variance_percent"]),
            "monthOverMonthGrowthPercent": float(row["revenue_growth_percent"]),
            "yearOverYearGrowthPercent": float(row["year_over_year_growth_percent"]),
            "riskScore": float(row["risk_score"]),
            "riskLevel": str(row["risk_level"]),
            "reason": str(row["reason"]),
            "bdm": str(row["bdm"]),
            "practiceHead": str(row["practice_head"]),
            "geoHead": str(row["geo_head"]),
            "vertical": str(row["vertical"]),
            "horizontal": str(row["horizontal"]),
            "entity": str(row.get("entity") or ""),
            "msps": str(row["ms_ps"]),
        }
        for _, row in risk_frame.head(200).iterrows()
    ]
    return {
        "rows": rows,
        "summary": {
            "low": int((frame["risk_level"] == "Low").sum()),
            "medium": int((frame["risk_level"] == "Medium").sum()),
            "high": int((frame["risk_level"] == "High").sum()),
            "criticalAnomalies": int((frame["anomaly_severity"] == "Critical").sum()),
        },
        "lastUpdated": _frame_last_updated(frame),
    }


def get_anomalies(filters: dict[str, Any] | None = None) -> dict[str, Any]:
    rows = _load_insight_logs(filters, insight_types=["Anomaly Alert"])
    return {
        "rows": rows,
        "count": len(rows),
        "lastUpdated": rows[0]["createdAt"] if rows else None,
    }


def get_predictions(filters: dict[str, Any] | None = None) -> dict[str, Any]:
    frame = _load_predictions_frame(filters)
    if frame.empty:
        return {"rows": [], "lastUpdated": None}

    frame = frame.sort_values(
        ["year", "month_sort_key", "prediction_level", "customer_name", "project_name"]
    )
    rows = [
        {
            "level": str(row["prediction_level"]),
            "customerName": str(row["customer_name"]),
            "projectName": str(row["project_name"]),
            "name": str(row.get("dimension_name") or _prediction_name(row)),
            "bdm": str(row["bdm"]),
            "practiceHead": str(row["practice_head"]),
            "geoHead": str(row["geo_head"]),
            "vertical": str(row["vertical"]),
            "horizontal": str(row["horizontal"]),
            "msps": str(row["ms_ps"]),
            "region": str(row.get("region") or ""),
            "salesRegion": str(row.get("sales_region") or ""),
            "month": str(row["month"]),
            "year": int(row["year"]),
            "quarter": str(row["quarter"]),
            "budget": float(row["budget_amount"]),
            "actual": float(row["actual_revenue"]),
            "previousMonthRevenue": float(row.get("previous_month_revenue") or 0.0),
            "previousYearRevenue": float(row.get("previous_year_revenue") or 0.0),
            "predictedRevenue": float(row["predicted_revenue"]),
            "lowerBound": float(row["lower_bound"]),
            "upperBound": float(row["upper_bound"]),
            "confidence": float(row["confidence_score"]),
            "predictionMethod": str(row.get("prediction_method") or row.get("model_used") or ""),
            "modelUsed": str(row["model_used"]),
            "modelVersion": str(row["model_version"]),
        }
        for _, row in frame.head(300).iterrows()
    ]
    return {"rows": rows, "lastUpdated": _frame_last_updated(frame)}


def get_insights(filters: dict[str, Any] | None = None) -> dict[str, Any]:
    rows = _load_insight_logs(filters, exclude_types=["Anomaly Alert"])
    return {"rows": rows, "count": len(rows)}


def get_trend_summary_rows(filters: dict[str, Any] | None = None) -> dict[str, Any]:
    frame = _load_enriched_summary_frame(filters)
    if frame.empty:
        return {"rows": [], "count": 0, "lastUpdated": None}

    ordered = frame.sort_values(
        ["year", "month_sort_key", "risk_score", "customer_name", "project_name"],
        ascending=[True, True, False, True, True],
    )
    rows = [
        {
            "customerName": str(row["customer_name"]),
            "projectName": str(row["project_name"]),
            "month": str(row["month"]),
            "year": int(row["year"]),
            "quarter": str(row["quarter"]),
            "budget": float(row["budget_amount"]),
            "actual": float(row["actual_revenue"]),
            "variance": float(row["budget_variance"]),
            "variancePercent": float(row["budget_variance_percent"]),
            "achievementPercent": float(row["budget_achievement_percent"]),
            "growthPercent": float(row["revenue_growth_percent"]),
            "previousMonthRevenue": float(row["previous_month_revenue"]),
            "previousYearRevenue": float(row["previous_year_revenue"]),
            "yearOverYearGrowthPercent": float(row["year_over_year_growth_percent"]),
            "marginPercent": float(row["margin_percent"]),
            "utilizationPercent": float(row["utilization_percent"]),
            "predictedRevenue": float(row["predicted_revenue"]),
            "predictionMethod": str(row.get("prediction_method") or ""),
            "predictionConfidence": float(row.get("prediction_confidence") or 0.0),
            "riskScore": float(row["risk_score"]),
            "riskLevel": str(row["risk_level"]),
            "riskReason": str(row.get("risk_reason") or ""),
            "anomalyFlag": bool(row["anomaly_flag"]),
            "anomalyReason": str(row["anomaly_reason"]),
            "bdm": str(row["bdm"]),
            "practiceHead": str(row["practice_head"]),
            "geoHead": str(row["geo_head"]),
            "vertical": str(row["vertical"]),
            "horizontal": str(row["horizontal"]),
            "msps": str(row["ms_ps"]),
            "region": str(row["region"]),
            "salesRegion": str(row["sales_region"]),
            "strategicAccount": str(row["strategic_account"]),
        }
        for _, row in ordered.iterrows()
    ]
    return {"rows": rows, "count": len(rows), "lastUpdated": _frame_last_updated(frame)}


def export_trend_excel(
    filters: dict[str, Any] | None = None,
    generated_by: str | None = None,
) -> tuple[str, bytes]:
    kpis_payload = get_trend_kpis(filters)
    monthly_payload = get_monthly_comparison(filters)
    budget_payload = get_budget_vs_actual(filters)
    yoy_payload = get_year_over_year(filters)
    risk_payload = get_risk_data(filters)
    summary_payload = get_trend_summary_rows(filters)
    anomalies_payload = get_anomalies(filters)
    predictions_payload = get_predictions(filters)
    insights_payload = get_insights(filters)

    workbook = Workbook()
    kpi_sheet = workbook.active
    kpi_sheet.title = "KPI Summary"
    _append_sheet_rows(kpi_sheet, kpis_payload["cards"])

    budget_sheet = workbook.create_sheet("Budget vs Actual")
    _append_sheet_rows(budget_sheet, monthly_payload["rows"])

    yoy_sheet = workbook.create_sheet("Year-over-Year")
    _append_sheet_rows(yoy_sheet, yoy_payload["rows"])

    risk_sheet = workbook.create_sheet("Risks")
    _append_sheet_rows(risk_sheet, risk_payload["rows"])

    anomalies_sheet = workbook.create_sheet("Anomalies")
    _append_sheet_rows(anomalies_sheet, anomalies_payload["rows"])

    predictions_sheet = workbook.create_sheet("Predictions")
    _append_sheet_rows(predictions_sheet, predictions_payload["rows"])

    insights_sheet = workbook.create_sheet("Insights")
    _append_sheet_rows(insights_sheet, insights_payload["rows"])

    detail_sheet = workbook.create_sheet("Detailed Data")
    _append_sheet_rows(detail_sheet, summary_payload["rows"])

    metadata_sheet = workbook.create_sheet("Export Metadata")
    metadata_sheet.append(["generated_by", sanitize_export_cell(generated_by or "system")])
    metadata_sheet.append(["generated_at", utc_now_iso()])
    metadata_sheet.append(["module", "trends"])
    metadata_sheet.append([
        "financial_year",
        str((filters or {}).get("financialYear") or (filters or {}).get("fy_year") or "all"),
    ])

    output = io.BytesIO()
    workbook.save(output)
    financial_year = str((filters or {}).get("financialYear") or (filters or {}).get("fy_year") or "all").strip() or "all"
    return (f"rapid-trends-{financial_year}.xlsx", output.getvalue())


def _refresh_trend_tables_for_year(financial_year: str, connection: Any) -> dict[str, Any]:
    budget_frame = _fetch_dataframe(
        connection,
        "select * from budget_data where fy_year = %s",
        (financial_year,),
    )
    budget_frame, budget_fallback_applied = _ensure_budget_frame_for_year(
        connection,
        budget_frame,
        financial_year,
    )
    actual_frame = _fetch_dataframe(
        connection,
        """
        select a.*
        from actual_revenue a
        join global_revenue_uploads u on u.id = a.uploaded_file_id
        where a.fy_year = %s
          and u.financial_year = %s
          and u.is_active = true
        """,
        (financial_year, financial_year),
    )

    trend_frame = _build_trend_summary_frame(budget_frame, actual_frame, financial_year)
    trend_frame = _attach_previous_year_metrics(trend_frame, connection, financial_year)
    trend_frame, anomalies = apply_anomaly_detection(trend_frame, financial_year)
    trend_frame = apply_risk_scores(trend_frame)
    predictions = build_predictions(trend_frame, financial_year)
    trend_frame = _attach_prediction_metrics(trend_frame, predictions)
    trend_frame = _finalize_summary_frame(trend_frame, financial_year)
    insights = build_insights(trend_frame, predictions, anomalies, financial_year)

    with connection.cursor() as cursor:
        cursor.execute("delete from trend_summary where fy_year = %s", (financial_year,))
        cursor.execute("delete from ml_predictions where fy_year = %s", (financial_year,))
        cursor.execute("delete from insight_logs where fy_year = %s", (financial_year,))
        if not trend_frame.empty:
            cursor.executemany(
                """
                insert into trend_summary (
                    customer_name,
                    project_name,
                    month,
                    year,
                    quarter,
                    bdm,
                    practice_head,
                    geo_head,
                    vertical,
                    horizontal,
                    entity,
                    ms_ps,
                    region,
                    sales_region,
                    strategic_account,
                    budget_amount,
                    actual_revenue,
                    budget_variance,
                    budget_variance_percent,
                    budget_achievement_percent,
                    previous_month_revenue,
                    previous_year_revenue,
                    revenue_growth_percent,
                    year_over_year_growth_percent,
                    margin_amount,
                    margin_percent,
                    utilization_percent,
                    predicted_revenue,
                    prediction_confidence,
                    prediction_method,
                    risk_score,
                    risk_level,
                    risk_reason,
                    anomaly_flag,
                    anomaly_reason,
                    anomaly_severity,
                    fy_year,
                    created_at,
                    updated_at
                )
                values (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::timestamptz, %s::timestamptz
                )
                """,
                [
                    (
                        str(row["customer_name"]),
                        str(row["project_name"]),
                        str(row["month"]),
                        int(row["year"]),
                        str(row["quarter"]),
                        str(row["bdm"]),
                        str(row["practice_head"]),
                        str(row["geo_head"]),
                        str(row["vertical"]),
                        str(row["horizontal"]),
                        str(row.get("entity") or ""),
                        str(row["ms_ps"]),
                        str(row["region"]),
                        str(row["sales_region"]),
                        str(row["strategic_account"]),
                        float(row["budget_amount"]),
                        float(row["actual_revenue"]),
                        float(row["budget_variance"]),
                        float(row["budget_variance_percent"]),
                        float(row["budget_achievement_percent"]),
                        float(row["previous_month_revenue"]),
                        float(row["previous_year_revenue"]),
                        float(row["revenue_growth_percent"]),
                        float(row["year_over_year_growth_percent"]),
                        float(row["margin_amount"]),
                        float(row["margin_percent"]),
                        float(row["utilization_percent"]),
                        float(row["predicted_revenue"]),
                        float(row["prediction_confidence"]),
                        str(row.get("prediction_method") or ""),
                        float(row["risk_score"]),
                        str(row["risk_level"]),
                        str(row.get("risk_reason") or ""),
                        bool(row["anomaly_flag"]),
                        str(row["anomaly_reason"]),
                        str(row["anomaly_severity"]),
                        financial_year,
                        row["created_at"],
                        row["updated_at"],
                    )
                    for _, row in trend_frame.iterrows()
                ],
            )
        if predictions:
            cursor.executemany(
                """
                insert into ml_predictions (
                    prediction_level,
                    customer_name,
                    project_name,
                    bdm,
                    practice_head,
                    geo_head,
                    vertical,
                    horizontal,
                    ms_ps,
                    dimension_name,
                    region,
                    sales_region,
                    month,
                    year,
                    quarter,
                    budget_amount,
                    actual_revenue,
                    previous_month_revenue,
                    previous_year_revenue,
                    predicted_revenue,
                    lower_bound,
                    upper_bound,
                    confidence_score,
                    prediction_method,
                    model_used,
                    model_version,
                    fy_year,
                    created_at
                )
                values (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::timestamptz
                )
                """,
                [
                    (
                        str(row.get("prediction_level") or ""),
                        str(row.get("customer_name") or ""),
                        str(row.get("project_name") or ""),
                        str(row.get("bdm") or ""),
                        str(row.get("practice_head") or ""),
                        str(row.get("geo_head") or ""),
                        str(row.get("vertical") or ""),
                        str(row.get("horizontal") or ""),
                        str(row.get("ms_ps") or ""),
                        str(row.get("dimension_name") or ""),
                        str(row.get("region") or ""),
                        str(row.get("sales_region") or ""),
                        str(row.get("month") or ""),
                        int(row.get("year") or 0),
                        str(row.get("quarter") or ""),
                        float(row.get("budget_amount") or 0.0),
                        float(row.get("actual_revenue") or 0.0),
                        float(row.get("previous_month_revenue") or 0.0),
                        float(row.get("previous_year_revenue") or 0.0),
                        float(row.get("predicted_revenue") or 0.0),
                        float(row.get("lower_bound") or 0.0),
                        float(row.get("upper_bound") or 0.0),
                        float(row.get("confidence_score") or 0.0),
                        str(row.get("prediction_method") or row.get("model_used") or ""),
                        str(row.get("model_used") or ""),
                        str(row.get("model_version") or ""),
                        financial_year,
                        utc_now_iso(),
                    )
                    for row in predictions
                ],
            )
        insight_rows = [*anomalies, *insights]
        if insight_rows:
            cursor.executemany(
                """
                insert into insight_logs (
                    insight_type,
                    severity,
                    title,
                    description,
                    recommendation,
                    dimension_type,
                    dimension_name,
                    customer_name,
                    project_name,
                    bdm,
                    practice_head,
                    geo_head,
                    vertical,
                    horizontal,
                    ms_ps,
                    month,
                    year,
                    metric_value,
                    comparison_value,
                    fy_year,
                    created_at
                )
                values (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::timestamptz
                )
                """,
                [
                    (
                        str(row.get("insight_type") or ""),
                        str(row.get("severity") or "Low"),
                        str(row.get("title") or ""),
                        str(row.get("description") or ""),
                        str(row.get("recommendation") or ""),
                        str(row.get("dimension_type") or ""),
                        str(row.get("dimension_name") or ""),
                        str(row.get("customer_name") or ""),
                        str(row.get("project_name") or ""),
                        str(row.get("bdm") or ""),
                        str(row.get("practice_head") or ""),
                        str(row.get("geo_head") or ""),
                        str(row.get("vertical") or ""),
                        str(row.get("horizontal") or ""),
                        str(row.get("ms_ps") or ""),
                        str(row.get("month") or ""),
                        int(row.get("year") or 0),
                        float(row.get("metric_value") or 0.0),
                        float(row.get("comparison_value") or 0.0),
                        financial_year,
                        utc_now_iso(),
                    )
                    for row in insight_rows
                ],
            )

    return {
        "rows_processed": len(trend_frame.index),
        "predictions_generated": len(predictions),
        "insights_generated": len(insights),
        "anomalies_detected": len(anomalies),
        "budget_fallback_applied": budget_fallback_applied,
    }


def _resolve_financial_years(connection: Any, financial_year: str | None) -> list[str]:
    normalized = str(financial_year or "").strip()
    if normalized:
        return [normalized]
    with connection.cursor() as cursor:
        cursor.execute(
            """
            select distinct fy_year
            from (
                select financial_year as fy_year from budget_uploads where is_active = true
                union
                select financial_year as fy_year from global_revenue_uploads where is_active = true
            ) years
            where coalesce(fy_year, '') <> ''
            order by fy_year
            """
        )
        return [
            str(row.get("fy_year") or "").strip()
            for row in cursor.fetchall()
            if str(row.get("fy_year") or "").strip()
        ]


def _fetch_dataframe(connection: Any, sql: str, params: tuple[Any, ...] | list[Any]) -> pd.DataFrame:
    with connection.cursor() as cursor:
        cursor.execute(sql, params)
        rows = cursor.fetchall()
    return pd.DataFrame(rows or [])


def _ensure_budget_frame_for_year(
    connection: Any,
    budget_frame: pd.DataFrame,
    financial_year: str,
) -> tuple[pd.DataFrame, bool]:
    if not budget_frame.empty:
        total_budget = float(_column_or_default(budget_frame, "budget_amount", 0).apply(_to_float).sum())
        if abs(total_budget) > 0.0:
            return budget_frame, False

    fallback = _load_budget_records_fallback_frame(connection, financial_year)
    if fallback.empty:
        return budget_frame, False
    return fallback, True


def _load_budget_records_fallback_frame(connection: Any, financial_year: str) -> pd.DataFrame:
    source = _fetch_dataframe(
        connection,
        """
        select
            r.customer_name,
            r.updated_customer,
            r.project_name,
            r.ms_ps,
            r.entity,
            r.row_us,
            r.strategic_account,
            r.practice_head,
            r.bdm,
            r.geo_head,
            r.vertical,
            r.horizontal,
            r.fy,
            r.q1,
            r.q2,
            r.q3,
            r.q4,
            r.apr_2026,
            r.may_2026,
            r.jun_2026,
            r.jul_2026,
            r.aug_2026,
            r.sep_2026,
            r.oct_2026,
            r.nov_2026,
            r.dec_2026,
            r.jan_2027,
            r.feb_2027,
            r.mar_2027,
            r.raw_payload
        from budget_records r
        join budget_uploads u on u.id = r.upload_id
        where u.financial_year = %s
          and u.is_active = true
          and r.financial_year = %s
        order by r.id asc
        """,
        (financial_year, financial_year),
    )
    if source.empty:
        return pd.DataFrame()

    start_year, end_year = parse_financial_year(financial_year)
    month_year_map = {
        month: (end_year if month in {"Jan", "Feb", "Mar"} else start_year)
        for month in MONTH_ORDER
    }
    rows: list[dict[str, Any]] = []
    for _, raw in source.iterrows():
        payload = raw.get("raw_payload") if isinstance(raw.get("raw_payload"), dict) else {}
        customer_name = _fallback_text(raw, payload, "updated_customer", ["Updated Customer", "Updated Customer Name"])
        if not customer_name:
            customer_name = _fallback_text(raw, payload, "customer_name", ["Customer Name", "Customer name", "Customer"])
        project_name = _fallback_text(raw, payload, "project_name", ["Project Name", "Project"])
        ms_ps = _fallback_text(raw, payload, "ms_ps", ["MS/PS", "PS/MS", "PS/MS budget", "MS/PS budget"])
        entity = _fallback_text(raw, payload, "entity", ["Entity", "Company"])
        row_us = _fallback_text(raw, payload, "row_us", ["ROW/US", "Region", "Region summary", "Region Summary"])
        strategic_account = _fallback_text(raw, payload, "strategic_account", ["Strategic Account"])
        practice_head = _fallback_text(raw, payload, "practice_head", ["Practice Head"])
        bdm = _fallback_text(raw, payload, "bdm", ["BDM"])
        geo_head = _fallback_text(raw, payload, "geo_head", ["Geo Head", "GeoHead", "Geo head"])
        vertical = _fallback_text(raw, payload, "vertical", ["Vertical"])
        horizontal = _fallback_text(raw, payload, "horizontal", ["Horizontal"])

        month_values = {
            month: _to_float(raw.get(column))
            for month, column in BUDGET_MONTH_FIELD_BY_NAME.items()
        }
        quarter_values = {
            "Q1": _to_float(raw.get("q1")),
            "Q2": _to_float(raw.get("q2")),
            "Q3": _to_float(raw.get("q3")),
            "Q4": _to_float(raw.get("q4")),
        }
        fy_total = _to_float(raw.get("fy"))
        monthly_total = sum(month_values.values())

        if abs(monthly_total) <= 0.0:
            for quarter, quarter_months in QUARTER_MONTHS.items():
                quarter_total = quarter_values.get(quarter, 0.0)
                if abs(quarter_total) <= 0.0:
                    continue
                if all(abs(month_values[month]) <= 0.0 for month in quarter_months):
                    split = quarter_total / float(len(quarter_months))
                    for month in quarter_months:
                        month_values[month] = split

            monthly_total = sum(month_values.values())
            if abs(monthly_total) <= 0.0 and abs(fy_total) > 0.0:
                split = fy_total / 12.0
                for month in MONTH_ORDER:
                    month_values[month] = split
            elif abs(fy_total) > 0.0 and monthly_total < fy_total:
                remaining = fy_total - monthly_total
                zero_months = [month for month in MONTH_ORDER if abs(month_values[month]) <= 0.0]
                if remaining > 0.0 and zero_months:
                    split = remaining / float(len(zero_months))
                    for month in zero_months:
                        month_values[month] += split

        for month in MONTH_ORDER:
            budget_amount = _to_float(month_values.get(month))
            if abs(budget_amount) <= 0.0:
                continue
            rows.append(
                {
                    "customer_name": customer_name,
                    "project_name": project_name,
                    "month": month,
                    "quarter": QUARTER_BY_MONTH.get(month, ""),
                    "year": month_year_map[month],
                    "budget_amount": budget_amount,
                    "bdm": bdm,
                    "practice_head": practice_head,
                    "geo_head": geo_head,
                    "vertical": vertical,
                    "horizontal": horizontal,
                    "entity": entity,
                    "ms_ps": ms_ps,
                    "strategic_account": strategic_account,
                    "row_us": row_us,
                    "fy_year": financial_year,
                }
            )

    return pd.DataFrame(rows)


def _fallback_text(
    row: pd.Series,
    payload: dict[str, Any],
    row_key: str,
    payload_keys: list[str],
) -> str:
    direct = _clean_text(row.get(row_key))
    if direct:
        return direct
    for key in payload_keys:
        candidate = _clean_text(payload.get(key))
        if candidate:
            return candidate
    return ""


def _build_trend_summary_frame(
    budget_frame: pd.DataFrame,
    actual_frame: pd.DataFrame,
    financial_year: str,
) -> pd.DataFrame:
    budget_grouped = _group_budget_frame(budget_frame, financial_year)
    actual_grouped = _group_actual_frame(actual_frame, financial_year)
    merged = _merge_budget_actual_frames(budget_grouped, actual_grouped, financial_year)
    if merged.empty:
        return merged

    merged["budget_amount"] = _column_or_default(merged, "budget_amount", 0.0).apply(_normalize_currency_amount)
    merged["actual_revenue"] = _column_or_default(merged, "actual_revenue", 0.0).apply(_normalize_currency_amount)
    merged["budget_variance"] = merged["actual_revenue"] - merged["budget_amount"]
    merged["budget_variance_percent"] = merged.apply(
        lambda row: _safe_budget_percent(row["budget_variance"], row["budget_amount"]),
        axis=1,
    )
    merged["budget_achievement_percent"] = merged.apply(
        lambda row: _safe_budget_percent(row["actual_revenue"], row["budget_amount"]),
        axis=1,
    )
    merged = merged.sort_values(["customer_name", "project_name", "month_sort_key"]).reset_index(drop=True)
    merged["previous_month_revenue"] = (
        merged.groupby(["customer_name", "project_name", "ms_ps"], dropna=False)["actual_revenue"]
        .shift(1)
        .fillna(0.0)
    )
    merged["revenue_growth_percent"] = merged.apply(
        lambda row: safe_pct(
            row["actual_revenue"] - row["previous_month_revenue"],
            row["previous_month_revenue"],
        )
        if float(row["previous_month_revenue"] or 0.0) > 0
        else 0.0,
        axis=1,
    )
    merged["margin_amount"] = (
        merged["actual_revenue"] - merged["expenses"] - merged["portal_fees"] - merged["tax"]
    )
    merged["margin_percent"] = merged.apply(
        lambda row: safe_pct(row["margin_amount"], row["actual_revenue"]),
        axis=1,
    )
    merged["utilization_percent"] = merged.apply(
        lambda row: safe_pct(row["billable_actual_hrs"], row["actual_hours"]),
        axis=1,
    )
    merged["risk_score"] = 0.0
    merged["risk_level"] = "Low"
    merged["anomaly_flag"] = False
    merged["anomaly_reason"] = ""
    merged["anomaly_severity"] = "Low"
    return merged


def _attach_previous_year_metrics(
    frame: pd.DataFrame,
    connection: Any,
    financial_year: str,
) -> pd.DataFrame:
    if frame.empty:
        frame["previous_year_revenue"] = []
        frame["year_over_year_growth_percent"] = []
        return frame

    enriched = frame.copy()
    previous_years = sorted({int(year) - 1 for year in enriched["year"].tolist() if int(year or 0) > 0})
    if not previous_years:
        enriched["previous_year_revenue"] = 0.0
        enriched["year_over_year_growth_percent"] = 0.0
        return enriched

    history_frame = _fetch_dataframe(
        connection,
        """
        select
            a.customer_name,
            a.project_name,
            a.ms_ps,
            a.month,
            a.year,
            a.actual_revenue_value
        from actual_revenue a
        join global_revenue_uploads u on u.id = a.uploaded_file_id
        where a.year = any(%s)
          and u.is_active = true
        """,
        (previous_years,),
    )
    if history_frame.empty:
        enriched["previous_year_revenue"] = 0.0
        enriched["year_over_year_growth_percent"] = 0.0
        return enriched

    history = history_frame.copy()
    history["customer_name"] = _column_or_default(history, "customer_name", "").apply(_clean_text)
    history["project_name"] = _column_or_default(history, "project_name", "").apply(_clean_text)
    history["month"] = _column_or_default(history, "month", "").apply(_clean_text)
    history["year"] = _column_or_default(history, "year", 0).apply(_to_int)
    history["actual_revenue"] = _column_or_default(history, "actual_revenue_value", 0).apply(_to_float)
    history["customer_key"] = history["customer_name"].apply(_normalize_key)
    history["project_key"] = history["project_name"].apply(_normalize_key)
    history["ms_ps"] = _column_or_default(history, "ms_ps", "").apply(_clean_text)

    exact_history = (
        history.groupby(["customer_key", "project_key", "ms_ps", "month", "year"], dropna=False)["actual_revenue"]
        .sum()
        .reset_index()
        .rename(columns={"year": "previous_year", "actual_revenue": "previous_year_revenue_exact"})
    )
    fallback_history = (
        history.groupby(["customer_key", "ms_ps", "month", "year"], dropna=False)
        .agg(
            previous_year_revenue_customer=("actual_revenue", "sum"),
            previous_project_count=("project_key", "nunique"),
        )
        .reset_index()
        .rename(columns={"year": "previous_year"})
    )

    enriched["previous_year"] = enriched["year"].apply(_to_int) - 1
    enriched = enriched.merge(
        exact_history,
        on=["customer_key", "project_key", "ms_ps", "month", "previous_year"],
        how="left",
    )
    enriched = enriched.merge(
        fallback_history,
        on=["customer_key", "ms_ps", "month", "previous_year"],
        how="left",
    )
    enriched["previous_year_revenue"] = enriched["previous_year_revenue_exact"].fillna(
        enriched.apply(
            lambda row: row["previous_year_revenue_customer"]
            if float(row.get("previous_project_count") or 0) == 1
            else 0.0,
            axis=1,
        )
    )
    enriched["year_over_year_growth_percent"] = enriched.apply(
        lambda row: safe_pct(
            float(row["actual_revenue"]) - float(row["previous_year_revenue"]),
            float(row["previous_year_revenue"]),
        )
        if float(row["previous_year_revenue"]) > 0
        else 0.0,
        axis=1,
    )
    return enriched.drop(
        columns=[
            "previous_year",
            "previous_year_revenue_exact",
            "previous_year_revenue_customer",
            "previous_project_count",
        ],
        errors="ignore",
    )


def _attach_prediction_metrics(
    frame: pd.DataFrame,
    predictions: list[dict[str, Any]],
) -> pd.DataFrame:
    if frame.empty:
        frame["predicted_revenue"] = []
        frame["prediction_confidence"] = []
        frame["prediction_method"] = []
        return frame
    if not predictions:
        enriched = frame.copy()
        enriched["predicted_revenue"] = 0.0
        enriched["prediction_confidence"] = 0.0
        enriched["prediction_method"] = ""
        return enriched

    prediction_frame = pd.DataFrame(predictions)
    project_predictions = prediction_frame[prediction_frame["prediction_level"] == "project"].copy()
    customer_predictions = prediction_frame[prediction_frame["prediction_level"] == "customer"].copy()

    enriched = frame.copy()
    enriched["customer_key"] = enriched["customer_name"].apply(_normalize_key)
    enriched["project_key"] = enriched["project_name"].apply(_normalize_key)

    for column in ("customer_name", "project_name", "month", "prediction_level"):
        if column in project_predictions:
            project_predictions[column] = project_predictions[column].fillna("").apply(_clean_text)
        if column in customer_predictions:
            customer_predictions[column] = customer_predictions[column].fillna("").apply(_clean_text)
    for column in ("year",):
        if column in project_predictions:
            project_predictions[column] = project_predictions[column].apply(_to_int)
        if column in customer_predictions:
            customer_predictions[column] = customer_predictions[column].apply(_to_int)

    if not project_predictions.empty:
        project_predictions["customer_key"] = project_predictions["customer_name"].apply(_normalize_key)
        project_predictions["project_key"] = project_predictions["project_name"].apply(_normalize_key)
        project_lookup = project_predictions[
            [
                "customer_key",
                "project_key",
                "month",
                "year",
                "predicted_revenue",
                "confidence_score",
                "prediction_method",
            ]
        ].rename(
            columns={
                "predicted_revenue": "predicted_revenue_exact",
                "confidence_score": "prediction_confidence_exact",
                "prediction_method": "prediction_method_exact",
            }
        )
        enriched = enriched.merge(
            project_lookup,
            on=["customer_key", "project_key", "month", "year"],
            how="left",
        )

    if not customer_predictions.empty:
        customer_predictions["customer_key"] = customer_predictions["customer_name"].apply(_normalize_key)
        customer_lookup = customer_predictions[
            [
                "customer_key",
                "month",
                "year",
                "predicted_revenue",
                "confidence_score",
                "prediction_method",
            ]
        ].rename(
            columns={
                "predicted_revenue": "predicted_revenue_customer",
                "confidence_score": "prediction_confidence_customer",
                "prediction_method": "prediction_method_customer",
            }
        )
        enriched = enriched.merge(
            customer_lookup,
            on=["customer_key", "month", "year"],
            how="left",
        )

    enriched["predicted_revenue"] = _coalesce_columns(
        enriched,
        ["predicted_revenue_exact", "predicted_revenue_customer"],
        0.0,
    ).apply(_to_float)
    enriched["prediction_confidence"] = _coalesce_columns(
        enriched,
        ["prediction_confidence_exact", "prediction_confidence_customer"],
        0.0,
    ).apply(_to_float)
    enriched["prediction_method"] = _coalesce_columns(
        enriched,
        ["prediction_method_exact", "prediction_method_customer"],
        "",
        treat_blank_as_missing=True,
    ).apply(_clean_text)
    return enriched.drop(
        columns=[
            "predicted_revenue_exact",
            "prediction_confidence_exact",
            "prediction_method_exact",
            "predicted_revenue_customer",
            "prediction_confidence_customer",
            "prediction_method_customer",
        ],
        errors="ignore",
    )


def _group_budget_frame(frame: pd.DataFrame, financial_year: str) -> pd.DataFrame:
    if frame.empty:
        return pd.DataFrame()
    budget = frame.copy()
    budget["customer_name"] = _column_or_default(budget, "customer_name", "").apply(_clean_text)
    budget["project_name"] = _column_or_default(budget, "project_name", "").apply(_clean_text)
    budget["month"] = _column_or_default(budget, "month", "").apply(_clean_text)
    budget["quarter"] = _column_or_default(budget, "quarter", "").apply(_clean_text)
    budget["year"] = _column_or_default(budget, "year", 0).apply(_to_int)
    budget["budget_amount"] = _column_or_default(budget, "budget_amount", 0).apply(_to_float)
    budget["customer_key"] = budget["customer_name"].apply(_normalize_key)
    budget["project_key"] = budget["project_name"].apply(_normalize_key)
    for column in [
        "bdm",
        "practice_head",
        "geo_head",
        "vertical",
        "horizontal",
        "entity",
        "ms_ps",
        "strategic_account",
        "row_us",
    ]:
        budget[column] = _column_or_default(budget, column, "").apply(
            lambda value, column_name=column: _normalize_summary_dimension_value("region", value)
            if column_name == "row_us"
            else _normalize_summary_dimension_value(column_name, value)
            if column_name in {"vertical", "horizontal", "ms_ps", "strategic_account"}
            else _clean_text(value)
        )
    aggregated = (
        budget.groupby(
            ["customer_key", "project_key", "ms_ps", "month", "year"],
            dropna=False,
        )
        .agg(
            budget_amount=("budget_amount", "sum"),
            customer_name=("customer_name", _first_non_blank),
            project_name=("project_name", _first_non_blank),
            quarter=("quarter", _first_non_blank),
            bdm=("bdm", _first_non_blank),
            practice_head=("practice_head", _first_non_blank),
            geo_head=("geo_head", _first_non_blank),
            vertical=("vertical", _first_non_blank),
            horizontal=("horizontal", _first_non_blank),
            entity=("entity", _first_non_blank),
            ms_ps=("ms_ps", _first_non_blank),
            strategic_account=("strategic_account", _first_non_blank),
            region=("row_us", _first_non_blank),
        )
        .reset_index()
    )
    aggregated["sales_region"] = ""
    aggregated["fy_year"] = financial_year
    aggregated["month_sort_key"] = aggregated.apply(
        lambda row: month_sort_key(str(row["month"]), int(row["year"] or 0)),
        axis=1,
    )
    return aggregated


def _group_actual_frame(frame: pd.DataFrame, financial_year: str) -> pd.DataFrame:
    if frame.empty:
        return pd.DataFrame()
    actual = frame.copy()
    actual["customer_name"] = _column_or_default(actual, "customer_name", "").apply(_clean_text)
    actual["project_name"] = _column_or_default(actual, "project_name", "").apply(_clean_text)
    actual["month"] = _column_or_default(actual, "month", "").apply(_clean_text)
    actual["quarter"] = _column_or_default(actual, "quarter", "").apply(_clean_text)
    actual["year"] = _column_or_default(actual, "year", 0).apply(_to_int)
    actual["actual_revenue"] = _column_or_default(actual, "actual_revenue_value", 0).apply(_to_float)
    actual["expenses"] = _column_or_default(actual, "expenses", 0).apply(_to_float)
    actual["portal_fees"] = _column_or_default(actual, "portal_fees", 0).apply(_to_float)
    actual["tax"] = _column_or_default(actual, "tax", 0).apply(_to_float)
    actual["billable_actual_hrs"] = _column_or_default(actual, "billable_actual_hrs", 0).apply(_to_float)
    actual["actual_hours"] = _column_or_default(actual, "actual_hours", 0).apply(_to_float)
    actual["customer_key"] = actual["customer_name"].apply(_normalize_key)
    actual["project_key"] = actual["project_name"].apply(_normalize_key)
    for column in [
        "bdm",
        "practice_head",
        "geo_head",
        "vertical",
        "horizontal",
        "entity",
        "ms_ps",
        "region",
        "sales_region",
    ]:
        source_column = "company" if column == "entity" else column
        actual[column] = _column_or_default(actual, source_column, "").apply(
            lambda value, column_name=column: _normalize_summary_dimension_value(column_name, value)
            if column_name in {"vertical", "horizontal", "ms_ps", "region", "sales_region"}
            else _clean_text(value)
        )
    aggregated = (
        actual.groupby(
            ["customer_key", "project_key", "ms_ps", "month", "year"],
            dropna=False,
        )
        .agg(
            actual_revenue=("actual_revenue", "sum"),
            expenses=("expenses", "sum"),
            portal_fees=("portal_fees", "sum"),
            tax=("tax", "sum"),
            billable_actual_hrs=("billable_actual_hrs", "sum"),
            actual_hours=("actual_hours", "sum"),
            customer_name=("customer_name", _first_non_blank),
            project_name=("project_name", _first_non_blank),
            quarter=("quarter", _first_non_blank),
            bdm=("bdm", _first_non_blank),
            practice_head=("practice_head", _first_non_blank),
            geo_head=("geo_head", _first_non_blank),
            vertical=("vertical", _first_non_blank),
            horizontal=("horizontal", _first_non_blank),
            entity=("entity", _first_non_blank),
            ms_ps=("ms_ps", _first_non_blank),
            region=("region", _first_non_blank),
            sales_region=("sales_region", _first_non_blank),
        )
        .reset_index()
    )
    aggregated["strategic_account"] = ""
    aggregated["fy_year"] = financial_year
    aggregated["month_sort_key"] = aggregated.apply(
        lambda row: month_sort_key(str(row["month"]), int(row["year"] or 0)),
        axis=1,
    )
    return aggregated


def _merge_budget_actual_frames(
    budget_grouped: pd.DataFrame,
    actual_grouped: pd.DataFrame,
    financial_year: str,
) -> pd.DataFrame:
    if budget_grouped.empty and actual_grouped.empty:
        return pd.DataFrame()
    if budget_grouped.empty:
        actual_only = actual_grouped.copy()
        actual_only["budget_amount"] = 0.0
        actual_only["strategic_account"] = _column_or_default(actual_only, "strategic_account", "").fillna("")
        return _coalesce_merged_frame(actual_only, financial_year)
    if actual_grouped.empty:
        budget_only = budget_grouped.copy()
        budget_only["actual_revenue"] = 0.0
        budget_only["expenses"] = 0.0
        budget_only["portal_fees"] = 0.0
        budget_only["tax"] = 0.0
        budget_only["billable_actual_hrs"] = 0.0
        budget_only["actual_hours"] = 0.0
        budget_only["sales_region"] = _column_or_default(budget_only, "sales_region", "").fillna("")
        return _coalesce_merged_frame(budget_only, financial_year)

    left = budget_grouped.copy().reset_index(drop=True)
    left["budget_row_id"] = left.index.astype(int)
    right = actual_grouped.copy().reset_index(drop=True)
    right["actual_row_id"] = right.index.astype(int)

    exact = left.merge(
        right,
        on=["customer_key", "project_key", "ms_ps", "month", "year"],
        how="outer",
        suffixes=("_budget", "_actual"),
        indicator=True,
    )
    exact_matched = exact.loc[exact["_merge"] == "both"].copy()

    matched_budget_ids = set(
        exact_matched.loc[exact_matched["budget_row_id"].notna(), "budget_row_id"].astype(int).tolist()
    )
    matched_actual_ids = set(
        exact_matched.loc[exact_matched["actual_row_id"].notna(), "actual_row_id"].astype(int).tolist()
    )
    budget_unmatched = left.loc[~left["budget_row_id"].isin(matched_budget_ids)].copy()
    actual_unmatched = right.loc[~right["actual_row_id"].isin(matched_actual_ids)].copy()

    fallback_rows = pd.DataFrame()
    if not budget_unmatched.empty and not actual_unmatched.empty:
        budget_counts = budget_unmatched.groupby(["customer_key", "ms_ps", "month", "year"], dropna=False).size().reset_index(name="budget_count")
        actual_counts = actual_unmatched.groupby(["customer_key", "ms_ps", "month", "year"], dropna=False).size().reset_index(name="actual_count")
        unique_pairs = budget_counts.merge(
            actual_counts,
            on=["customer_key", "ms_ps", "month", "year"],
            how="inner",
        )
        unique_pairs = unique_pairs[
            (unique_pairs["budget_count"] == 1) & (unique_pairs["actual_count"] == 1)
        ]
        if not unique_pairs.empty:
            budget_unique = budget_unmatched.merge(
                unique_pairs[["customer_key", "ms_ps", "month", "year"]],
                on=["customer_key", "ms_ps", "month", "year"],
                how="inner",
            )
            actual_unique = actual_unmatched.merge(
                unique_pairs[["customer_key", "ms_ps", "month", "year"]],
                on=["customer_key", "ms_ps", "month", "year"],
                how="inner",
            )
            fallback_rows = budget_unique.merge(
                actual_unique,
                on=["customer_key", "ms_ps", "month", "year"],
                how="inner",
                suffixes=("_budget", "_actual"),
            )
            used_budget_ids = set(fallback_rows["budget_row_id"].astype(int).tolist())
            used_actual_ids = set(fallback_rows["actual_row_id"].astype(int).tolist())
            budget_unmatched = budget_unmatched.loc[~budget_unmatched["budget_row_id"].isin(used_budget_ids)].copy()
            actual_unmatched = actual_unmatched.loc[~actual_unmatched["actual_row_id"].isin(used_actual_ids)].copy()

    remaining_budget = budget_unmatched.copy()
    if not remaining_budget.empty:
        for column in [
            "actual_revenue",
            "expenses",
            "portal_fees",
            "tax",
            "billable_actual_hrs",
            "actual_hours",
        ]:
            remaining_budget[column] = 0.0
        for column in [
            "quarter_actual",
            "bdm_actual",
            "practice_head_actual",
            "geo_head_actual",
            "vertical_actual",
            "horizontal_actual",
            "entity_actual",
            "ms_ps_actual",
            "region_actual",
            "sales_region_actual",
            "project_name_actual",
            "customer_name_actual",
        ]:
            remaining_budget[column] = ""

    remaining_actual = actual_unmatched.copy()
    if not remaining_actual.empty:
        remaining_actual["budget_amount"] = 0.0
        for column in [
            "quarter_budget",
            "bdm_budget",
            "practice_head_budget",
            "geo_head_budget",
            "vertical_budget",
            "horizontal_budget",
            "entity_budget",
            "ms_ps_budget",
            "strategic_account_budget",
            "region_budget",
            "sales_region_budget",
            "project_name_budget",
            "customer_name_budget",
        ]:
            remaining_actual[column] = ""

    merged_rows = pd.concat(
        [exact_matched, fallback_rows, remaining_budget, remaining_actual],
        ignore_index=True,
        sort=False,
    )
    return _coalesce_merged_frame(merged_rows, financial_year)


def _coalesce_merged_frame(frame: pd.DataFrame, financial_year: str) -> pd.DataFrame:
    combined = frame.copy()
    combined["customer_name"] = _coalesce_columns(
        combined,
        ["customer_name", "customer_name_budget", "customer_name_actual"],
        "",
        treat_blank_as_missing=True,
    )
    combined["project_name"] = _coalesce_columns(
        combined,
        ["project_name", "project_name_budget", "project_name_actual"],
        "",
        treat_blank_as_missing=True,
    )
    combined["quarter"] = _coalesce_columns(
        combined,
        ["quarter", "quarter_budget", "quarter_actual"],
        "",
        treat_blank_as_missing=True,
    )
    for base in [
        "bdm",
        "practice_head",
        "geo_head",
        "vertical",
        "horizontal",
        "entity",
        "ms_ps",
        "region",
        "sales_region",
        "strategic_account",
    ]:
        budget_column = f"{base}_budget"
        actual_column = f"{base}_actual"
        combined[base] = _coalesce_columns(
            combined,
            [base, budget_column, actual_column],
            "",
            treat_blank_as_missing=True,
        )
    combined["budget_amount"] = _coalesce_columns(
        combined,
        ["budget_amount", "budget_amount_budget"],
        0.0,
    )
    combined["actual_revenue"] = _coalesce_columns(
        combined,
        ["actual_revenue", "actual_revenue_actual"],
        0.0,
    )
    for numeric_column in [
        "budget_amount",
        "actual_revenue",
        "expenses",
        "portal_fees",
        "tax",
        "billable_actual_hrs",
        "actual_hours",
    ]:
        combined[numeric_column] = _column_or_default(combined, numeric_column, 0.0).fillna(0.0).apply(_to_float)
    combined["customer_name"] = combined["customer_name"].fillna("").apply(_clean_text)
    combined["project_name"] = combined["project_name"].fillna("").apply(_clean_text)
    combined["month"] = _column_or_default(combined, "month", "").fillna("").apply(_clean_text)
    combined["quarter"] = combined["quarter"].fillna("").apply(_clean_text)
    combined["year"] = _column_or_default(combined, "year", 0).fillna(0).apply(_to_int)
    combined["customer_key"] = combined["customer_name"].apply(_normalize_key)
    combined["project_key"] = combined["project_name"].apply(_normalize_key)
    combined["fy_year"] = financial_year
    combined["month_sort_key"] = combined.apply(
        lambda row: month_sort_key(str(row["month"]), int(row["year"] or 0)),
        axis=1,
    )
    return combined


def _finalize_summary_frame(frame: pd.DataFrame, financial_year: str) -> pd.DataFrame:
    if frame.empty:
        return frame
    finalized = frame.copy()
    now_iso = utc_now_iso()
    finalized["fy_year"] = financial_year
    finalized["created_at"] = now_iso
    finalized["updated_at"] = now_iso
    for text_column in SUMMARY_TEXT_FIELDS:
        finalized[text_column] = _column_or_default(finalized, text_column, "").fillna("").apply(
            lambda value, column_name=text_column: _normalize_summary_dimension_value(column_name, value)
            if column_name in {"vertical", "horizontal", "ms_ps", "region", "sales_region", "strategic_account"}
            else _clean_text(value)
        )
    for numeric_column in SUMMARY_NUMERIC_FIELDS:
        finalized[numeric_column] = _column_or_default(finalized, numeric_column, 0.0).fillna(0.0).apply(_to_float)
    for amount_column in ["budget_amount", "actual_revenue", "budget_variance"]:
        finalized[amount_column] = _column_or_default(finalized, amount_column, 0.0).apply(_normalize_currency_amount)
    for percent_column in ["budget_variance_percent", "budget_achievement_percent"]:
        finalized[percent_column] = _column_or_default(finalized, percent_column, 0.0).apply(_clip_numeric_18_2)
    for numeric_column in SUMMARY_NUMERIC_FIELDS:
        finalized[numeric_column] = _column_or_default(finalized, numeric_column, 0.0).apply(_clip_numeric_18_2)
    finalized["anomaly_reason"] = _column_or_default(finalized, "anomaly_reason", "").fillna("").apply(_clean_text)
    finalized["anomaly_severity"] = _column_or_default(finalized, "anomaly_severity", "Low").fillna("Low").apply(_clean_text)
    finalized["anomaly_flag"] = _column_or_default(finalized, "anomaly_flag", False).fillna(False).astype(bool)
    finalized["risk_level"] = _column_or_default(finalized, "risk_level", "Low").fillna("Low").apply(_clean_text)
    finalized["month"] = _column_or_default(finalized, "month", "").fillna("").apply(_clean_text)
    finalized["quarter"] = _column_or_default(finalized, "quarter", "").fillna("").apply(_clean_text)
    finalized["year"] = _column_or_default(finalized, "year", 0).fillna(0).apply(_to_int)
    finalized["month_sort_key"] = finalized.apply(
        lambda row: month_sort_key(str(row["month"]), int(row["year"] or 0)),
        axis=1,
    )
    return finalized


def _load_summary_frame(filters: dict[str, Any] | None = None) -> pd.DataFrame:
    ensure_postgres_schema()
    where_sql, params = _build_where_clause(filters, alias="t")
    with open_database_connection(require=True) as connection:
        assert connection is not None
        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                select *
                from trend_summary t
                {where_sql}
                order by t.year asc, t.month asc, t.customer_name asc, t.project_name asc
                """,
                params,
            )
            rows = cursor.fetchall()
    return pd.DataFrame(rows or [])


def _prepare_summary_dataframe(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return frame
    prepared = frame.copy()
    for numeric_column in [
        *SUMMARY_NUMERIC_FIELDS,
        "expenses",
        "portal_fees",
        "tax",
        "billable_actual_hrs",
        "actual_hours",
    ]:
        if numeric_column in prepared:
            prepared[numeric_column] = prepared[numeric_column].apply(_to_float)
    prepared["year"] = _column_or_default(prepared, "year", 0).apply(_to_int)
    prepared["month_sort_key"] = prepared.apply(
        lambda row: month_sort_key(str(row["month"]), int(row["year"] or 0)),
        axis=1,
    )
    return prepared


def _load_enriched_summary_frame(filters: dict[str, Any] | None = None) -> pd.DataFrame:
    return _prepare_summary_dataframe(_load_summary_frame(filters))


def _load_predictions_frame(
    filters: dict[str, Any] | None = None,
    prediction_levels: list[str] | None = None,
) -> pd.DataFrame:
    ensure_postgres_schema()
    where_sql, params = _build_where_clause(filters, alias="p")
    if prediction_levels:
        placeholders = ", ".join(["%s"] * len(prediction_levels))
        where_sql = f"{where_sql} {'and' if where_sql else 'where'} p.prediction_level in ({placeholders})"
        params.extend(prediction_levels)
    with open_database_connection(require=True) as connection:
        assert connection is not None
        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                select *
                from ml_predictions p
                {where_sql}
                order by p.year asc, p.month asc, p.prediction_level asc
                """,
                params,
            )
            rows = cursor.fetchall()
    frame = pd.DataFrame(rows or [])
    if frame.empty:
        return frame
    frame["predicted_revenue"] = _column_or_default(frame, "predicted_revenue", 0).apply(_to_float)
    frame["budget_amount"] = _column_or_default(frame, "budget_amount", 0).apply(_to_float)
    frame["actual_revenue"] = _column_or_default(frame, "actual_revenue", 0).apply(_to_float)
    frame["previous_month_revenue"] = _column_or_default(frame, "previous_month_revenue", 0).apply(_to_float)
    frame["previous_year_revenue"] = _column_or_default(frame, "previous_year_revenue", 0).apply(_to_float)
    frame["confidence_score"] = _column_or_default(frame, "confidence_score", 0).apply(_to_float)
    frame["lower_bound"] = _column_or_default(frame, "lower_bound", 0).apply(_to_float)
    frame["upper_bound"] = _column_or_default(frame, "upper_bound", 0).apply(_to_float)
    frame["year"] = _column_or_default(frame, "year", 0).apply(_to_int)
    frame["month_sort_key"] = frame.apply(
        lambda row: month_sort_key(str(row["month"]), int(row["year"] or 0)),
        axis=1,
    )
    return frame


def _load_insight_logs(
    filters: dict[str, Any] | None = None,
    insight_types: list[str] | None = None,
    exclude_types: list[str] | None = None,
) -> list[dict[str, Any]]:
    ensure_postgres_schema()
    where_sql, params = _build_where_clause(filters, alias="i")
    if insight_types:
        placeholders = ", ".join(["%s"] * len(insight_types))
        where_sql = f"{where_sql} {'and' if where_sql else 'where'} i.insight_type in ({placeholders})"
        params.extend(insight_types)
    if exclude_types:
        placeholders = ", ".join(["%s"] * len(exclude_types))
        where_sql = f"{where_sql} {'and' if where_sql else 'where'} i.insight_type not in ({placeholders})"
        params.extend(exclude_types)
    with open_database_connection(require=True) as connection:
        assert connection is not None
        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                select *
                from insight_logs i
                {where_sql}
                order by i.created_at desc, i.year desc, i.month asc
                """,
                params,
            )
            rows = cursor.fetchall()
    output: list[dict[str, Any]] = []
    for row in rows:
        output.append(
            {
                "type": str(row.get("insight_type") or ""),
                "severity": str(row.get("severity") or "Low"),
                "title": str(row.get("title") or ""),
                "description": str(row.get("description") or ""),
                "recommendation": str(row.get("recommendation") or ""),
                "dimensionType": str(row.get("dimension_type") or ""),
                "dimensionName": str(row.get("dimension_name") or ""),
                "customerName": str(row.get("customer_name") or ""),
                "projectName": str(row.get("project_name") or ""),
                "bdm": str(row.get("bdm") or ""),
                "practiceHead": str(row.get("practice_head") or ""),
                "geoHead": str(row.get("geo_head") or ""),
                "vertical": str(row.get("vertical") or ""),
                "horizontal": str(row.get("horizontal") or ""),
                "msps": str(row.get("ms_ps") or ""),
                "month": str(row.get("month") or ""),
                "year": int(row.get("year") or 0),
                "metricValue": float(_to_float(row.get("metric_value"))),
                "comparisonValue": float(_to_float(row.get("comparison_value"))),
                "createdAt": _serialize_timestamp(row.get("created_at")),
            }
        )
    return output


def _build_where_clause(
    filters: dict[str, Any] | None,
    *,
    alias: str,
) -> tuple[str, list[Any]]:
    if not filters:
        return "", []
    clauses: list[str] = []
    params: list[Any] = []
    for filter_key, column_name in FILTER_COLUMN_MAP.items():
        if filter_key not in filters:
            continue
        raw_value = filters.get(filter_key)
        if column_name == "year":
            values = [_to_int(value) for value in _expand_filter_values(raw_value) if _to_int(value) > 0]
            if not values:
                continue
            if len(values) == 1:
                clauses.append(f"{alias}.{column_name} = %s")
                params.append(values[0])
            else:
                clauses.append(f"{alias}.{column_name} = any(%s)")
                params.append(values)
            continue

        values = [str(value).strip() for value in _expand_filter_values(raw_value) if str(value).strip()]
        if not values:
            continue
        if len(values) == 1:
            clauses.append(f"lower(coalesce({alias}.{column_name}, '')) = %s")
            params.append(values[0].lower())
        else:
            clauses.append(f"lower(coalesce({alias}.{column_name}, '')) = any(%s)")
            params.append([value.lower() for value in values])
    if not clauses:
        return "", []
    return f"where {' and '.join(clauses)}", params


def _expand_filter_values(value: Any) -> list[Any]:
    if value is None:
        return []
    if isinstance(value, list):
        output: list[Any] = []
        for item in value:
            output.extend(_expand_filter_values(item))
        return output
    if isinstance(value, tuple):
        output: list[Any] = []
        for item in value:
            output.extend(_expand_filter_values(item))
        return output
    text = str(value).strip()
    if not text:
        return []
    return [part.strip() for part in text.replace("|", ",").replace(";", ",").split(",") if part.strip()]


def _build_dimension_breakdown(frame: pd.DataFrame, dimension: str) -> list[dict[str, Any]]:
    if dimension not in frame:
        return []
    grouped = (
        frame.groupby(dimension, dropna=False)
        .agg({"budget_amount": "sum", "actual_revenue": "sum"})
        .reset_index()
    )
    grouped[dimension] = grouped[dimension].fillna("").apply(_clean_text)
    grouped["label"] = grouped[dimension].replace("", "Unassigned")
    grouped["variance"] = grouped["actual_revenue"] - grouped["budget_amount"]
    total_actual = float(grouped["actual_revenue"].sum())
    grouped["contribution_percent"] = grouped["actual_revenue"].apply(
        lambda value: safe_pct(value, total_actual) if total_actual > 0 else 0.0
    )
    return [
        {
            "label": str(row["label"]),
            "budget": float(row["budget_amount"]),
            "actual": float(row["actual_revenue"]),
            "variance": float(row["variance"]),
            "contributionPercent": float(row["contribution_percent"]),
        }
        for _, row in grouped.sort_values("actual_revenue", ascending=False).head(12).iterrows()
    ]


def _build_underperformer_rows(frame: pd.DataFrame) -> list[dict[str, Any]]:
    scoped = frame[
        [
            "customer_name",
            "project_name",
            "month",
            "year",
            "budget_amount",
            "actual_revenue",
            "budget_variance",
            "budget_variance_percent",
            "budget_achievement_percent",
            "bdm",
            "practice_head",
            "geo_head",
            "risk_level",
        ]
    ].copy()
    scoped = scoped[(scoped["budget_amount"] > 0) & (scoped["budget_variance"] < 0)]
    scoped = scoped.sort_values(["budget_variance_percent", "budget_variance"])
    return [
        {
            "customerName": str(row["customer_name"]),
            "projectName": str(row["project_name"]),
            "month": str(row["month"]),
            "year": int(row["year"]),
            "budget": float(row["budget_amount"]),
            "actual": float(row["actual_revenue"]),
            "variance": float(row["budget_variance"]),
            "variancePercent": float(row["budget_variance_percent"]),
            "achievementPercent": float(row["budget_achievement_percent"]),
            "bdm": str(row["bdm"]),
            "practiceHead": str(row["practice_head"]),
            "geoHead": str(row["geo_head"]),
            "riskLevel": str(row["risk_level"]),
        }
        for _, row in scoped.head(50).iterrows()
    ]


def _build_yoy_breakdown(frame: pd.DataFrame, dimension: str) -> list[dict[str, Any]]:
    if dimension not in frame:
        return []
    grouped = (
        frame.groupby(dimension, dropna=False)
        .agg({"actual_revenue": "sum", "previous_year_revenue": "sum"})
        .reset_index()
    )
    grouped[dimension] = grouped[dimension].fillna("").apply(_clean_text)
    grouped["label"] = grouped[dimension].replace("", "Unassigned")
    grouped["yoyVariance"] = grouped["actual_revenue"] - grouped["previous_year_revenue"]
    grouped["yoyGrowthPercent"] = grouped.apply(
        lambda row: safe_pct(row["yoyVariance"], row["previous_year_revenue"])
        if float(row["previous_year_revenue"]) > 0
        else 0.0,
        axis=1,
    )
    grouped["status"] = grouped.apply(
        lambda row: _growth_status(float(row["actual_revenue"]), float(row["previous_year_revenue"])),
        axis=1,
    )
    return [
        {
            "label": str(row["label"]),
            "currentYearRevenue": float(row["actual_revenue"]),
            "previousYearRevenue": float(row["previous_year_revenue"]),
            "yoyVariance": float(row["yoyVariance"]),
            "yoyGrowthPercent": float(row["yoyGrowthPercent"]),
            "status": str(row["status"]),
        }
        for _, row in grouped.sort_values("actual_revenue", ascending=False).head(12).iterrows()
    ]


def _growth_status(current_value: float, previous_value: float) -> str:
    if previous_value <= 0:
        return "No Previous Year Data"
    if current_value > previous_value:
        return "Growth"
    if current_value < previous_value:
        return "Decline"
    return "No Change"


def _previous_period_totals(frame: pd.DataFrame) -> dict[str, float]:
    grouped = (
        frame.groupby(["month", "year", "month_sort_key"], dropna=False)
        .agg(
            {
                "budget_amount": "sum",
                "actual_revenue": "sum",
                "budget_variance": "sum",
                "budget_achievement_percent": "mean",
                "revenue_growth_percent": "mean",
            }
        )
        .reset_index()
        .sort_values("month_sort_key")
    )
    if len(grouped.index) < 2:
        return {}
    row = grouped.iloc[-2]
    return {
        column: float(_to_float(row[column]))
        for column in grouped.columns
        if column not in {"month", "year", "month_sort_key"}
    }


def _latest_period_totals(frame: pd.DataFrame) -> dict[str, float]:
    grouped = (
        frame.groupby(["month", "year", "month_sort_key"], dropna=False)
        .agg(
            {
                "budget_amount": "sum",
                "actual_revenue": "sum",
                "budget_variance": "sum",
                "budget_achievement_percent": "mean",
                "revenue_growth_percent": "mean",
            }
        )
        .reset_index()
        .sort_values("month_sort_key")
    )
    if grouped.empty:
        return {}
    row = grouped.iloc[-1]
    return {
        column: float(_to_float(row[column]))
        for column in grouped.columns
        if column not in {"month", "year", "month_sort_key"}
    }


def _frame_last_updated(frame: pd.DataFrame) -> str | None:
    if frame.empty or "updated_at" not in frame:
        return None
    last_value = frame["updated_at"].dropna().max()
    if pd.isna(last_value):
        return None
    if isinstance(last_value, pd.Timestamp):
        return last_value.to_pydatetime().isoformat().replace("+00:00", "Z")
    return str(last_value)


def _resolve_risk_reason(row: pd.Series) -> str:
    explicit_reason = _clean_text(row.get("anomaly_reason"))
    if explicit_reason:
        return explicit_reason
    budget = float(row.get("budget_amount") or 0.0)
    actual = float(row.get("actual_revenue") or 0.0)
    variance_pct = float(row.get("budget_variance_percent") or 0.0)
    growth_pct = float(row.get("revenue_growth_percent") or 0.0)
    if budget > 0 and actual <= 0:
        return "Budget exists but actual revenue is missing."
    if variance_pct <= -15:
        return "Actual revenue is materially below budget."
    if growth_pct < 0:
        return "Revenue is declining month over month."
    return "Risk score is elevated due to performance softness."


def _prediction_name(row: pd.Series) -> str:
    for key in (
        "dimension_name",
        "project_name",
        "customer_name",
        "bdm",
        "practice_head",
        "geo_head",
        "vertical",
        "horizontal",
        "ms_ps",
        "region",
        "sales_region",
    ):
        value = _clean_text(row.get(key))
        if value:
            return value
    return str(row.get("prediction_level") or "overall").replace("_", " ").title()


def _kpi_card(
    title: str,
    value: float,
    current_value: float | None,
    previous_value: float | None,
    description: str,
    format_type: str,
) -> dict[str, Any]:
    delta = None
    delta_percent = None
    if current_value is not None and previous_value is not None:
        delta = float(current_value) - float(previous_value)
        delta_percent = safe_pct(delta, float(previous_value)) if float(previous_value) != 0 else 0.0
    status = "neutral"
    metric = delta if format_type == "currency" else delta_percent
    if metric is not None:
        if metric > 0:
            status = "positive"
        elif metric < 0:
            status = "negative"
    return {
        "title": title,
        "value": float(value),
        "currentValue": float(current_value) if current_value is not None else None,
        "previousValue": float(previous_value) if previous_value is not None else None,
        "delta": delta,
        "deltaPercent": delta_percent,
        "status": status,
        "description": description,
        "format": format_type,
    }


def _append_sheet_rows(worksheet: Any, rows: list[dict[str, Any]]) -> None:
    if not rows:
        worksheet.append(["No data"])
        return
    headers = list(rows[0].keys())
    worksheet.append(headers)
    for row in rows:
        worksheet.append([sanitize_export_cell(row.get(header)) for header in headers])


def _column_or_default(frame: pd.DataFrame, column: str, default: Any) -> pd.Series:
    if column in frame:
        return frame[column]
    return pd.Series([default] * len(frame.index), index=frame.index)


def _coalesce_columns(
    frame: pd.DataFrame,
    columns: list[str],
    default: Any,
    *,
    treat_blank_as_missing: bool = False,
) -> pd.Series:
    candidates = pd.concat(
        [_column_or_default(frame, column, default) for column in columns],
        axis=1,
        ignore_index=True,
    )
    if treat_blank_as_missing:
        candidates = candidates.replace(r"^\s*$", pd.NA, regex=True)
    merged = candidates.bfill(axis=1).iloc[:, 0]
    return merged.fillna(default)


def _first_non_blank(series: pd.Series) -> str:
    for value in series.tolist():
        text = _clean_text(value)
        if text:
            return text
    return ""


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value).strip()


def _canonical_dimension_title(value: Any) -> str:
    text = re.sub(r"\s+", " ", _clean_text(value))
    if not text:
        return ""
    parts = re.split(r"(\s+|/|-)", text)
    normalized_parts: list[str] = []
    for part in parts:
        if not part or part.isspace() or part in {"/", "-"}:
            normalized_parts.append(part)
            continue
        if part.isupper() and len(part) <= 4:
            normalized_parts.append(part)
            continue
        normalized_parts.append(part[:1].upper() + part[1:].lower())
    return "".join(normalized_parts).strip()


def _normalize_summary_dimension_value(column: str, value: Any) -> str:
    text = re.sub(r"\s+", " ", _clean_text(value))
    if not text:
        return ""

    upper_text = text.upper()
    if column == "ms_ps":
        if upper_text in {"MS", "MS/PS"}:
            return "MS"
        if upper_text in {"PS", "PS/MS"}:
            return "PS"
        return upper_text
    if column in {"region", "sales_region"}:
        if upper_text in US_REGION_ALIASES:
            return "US"
        if upper_text in ROW_REGION_ALIASES:
            return "ROW"
        return _canonical_dimension_title(text)
    if column == "strategic_account":
        if upper_text in {"YES", "Y", "TRUE", "1"}:
            return "Yes"
        if upper_text in {"NO", "N", "FALSE", "0"}:
            return "No"
        return _canonical_dimension_title(text)
    if column in {"vertical", "horizontal"}:
        return _canonical_dimension_title(text)
    return text


def _normalize_key(value: Any) -> str:
    return _clean_text(value).lower()


def _to_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value))
    except (TypeError, ValueError):
        return 0.0


def _to_int(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return 0


def _normalize_currency_amount(value: Any) -> float:
    amount = round(_to_float(value), 2)
    if abs(amount) < CURRENCY_ZERO_EPSILON:
        return 0.0
    return _clip_numeric_18_2(amount)


def _safe_budget_percent(numerator: Any, denominator: Any) -> float:
    denominator_value = _normalize_currency_amount(denominator)
    if abs(denominator_value) < MIN_BUDGET_DENOMINATOR_FOR_PERCENT:
        return 0.0
    return _clip_numeric_18_2(safe_pct(_to_float(numerator), denominator_value))


def _clip_numeric_18_2(value: Any) -> float:
    numeric_value = _to_float(value)
    if numeric_value > NUMERIC_18_2_ABS_MAX:
        return NUMERIC_18_2_ABS_MAX
    if numeric_value < -NUMERIC_18_2_ABS_MAX:
        return -NUMERIC_18_2_ABS_MAX
    return numeric_value


def _serialize_timestamp(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat().replace("+00:00", "Z")
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime().isoformat().replace("+00:00", "Z")
    text = str(value).strip()
    return text or None
