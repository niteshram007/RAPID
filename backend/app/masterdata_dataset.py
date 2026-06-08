from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook

try:
    import xlrd
except Exception:  # pragma: no cover
    xlrd = None  # type: ignore[assignment]

FieldKind = Literal["text", "numeric", "date"]
DatasetType = Literal["budget", "global_revenue", "forecast"]


@dataclass(frozen=True)
class MasterdataFieldSpec:
    key: str
    label: str
    kind: FieldKind
    aliases: tuple[str, ...] = ()


@dataclass
class MasterdataParsedRow:
    sheet_name: str
    row_number: int
    values: dict[str, Any]
    raw_payload: dict[str, Any]


@dataclass
class MasterdataParsedWorkbook:
    rows: list[MasterdataParsedRow]
    parsed_sheets: list[str]
    matched_columns: list[str]
    missing_required_columns: list[str]
    invalid_rows: list[dict[str, Any]]


MASTERDATA_DATASET_TYPES: tuple[DatasetType, ...] = (
    "budget",
    "global_revenue",
    "forecast",
)

MASTERDATA_FIELD_SPECS: tuple[MasterdataFieldSpec, ...] = (
    MasterdataFieldSpec(
        "customer_name",
        "Customer Name",
        "text",
        aliases=(
            "Customer",
            "Customer name",
        ),
    ),
    MasterdataFieldSpec(
        "updated_customer",
        "Updated Customer",
        "text",
        aliases=("Updated Customer Name",),
    ),
    MasterdataFieldSpec(
        "ms_ps",
        "MS/PS",
        "text",
        aliases=("PS/MS", "PS/MS budget", "MS/PS budget", "MS PS", "PS MS"),
    ),
    MasterdataFieldSpec("entity", "Entity", "text", aliases=("Company",)),
    MasterdataFieldSpec(
        "gr_entity",
        "GR Entity",
        "text",
        aliases=("Entity As per GR", "Group company", "Group Company"),
    ),
    MasterdataFieldSpec(
        "row_us",
        "ROW/US",
        "text",
        aliases=("Region", "Region 2", "Region summary", "Region Summary"),
    ),
    MasterdataFieldSpec("strategic_account", "Strategic Account", "text"),
    MasterdataFieldSpec(
        "resource_id",
        "Emp ID",
        "text",
        aliases=("Resource ID", "Emp ID", "Employee ID", "Employee Id"),
    ),
    MasterdataFieldSpec("resource_name", "Resource Name", "text"),
    MasterdataFieldSpec("deal_type", "Deal Type", "text", aliases=("Revenue type", "Revenue Type")),
    MasterdataFieldSpec("eeennn", "EEENNN", "text", aliases=("EEENNN", "EENNN")),
    MasterdataFieldSpec("bill_rate", "Bill Rate", "numeric", aliases=("Bill Rat",)),
    MasterdataFieldSpec("rate_type", "Rate Type", "text", aliases=("Rate Type", "RateType")),
    MasterdataFieldSpec(
        "billed_currency",
        "Billed currency",
        "text",
        aliases=("Billed Currency", "Currency"),
    ),
    MasterdataFieldSpec(
        "forex",
        "Forex",
        "numeric",
        aliases=("FX", "FX Rate", "Forex Rate", "Exchange Rate"),
    ),
    MasterdataFieldSpec(
        "type_of_projects",
        "Type of Projects",
        "text",
        aliases=("Type of Project", "Type of Projects"),
    ),
    MasterdataFieldSpec("billed_hours", "Billed Hours", "numeric"),
    MasterdataFieldSpec("billable_actual_hrs", "Billable Actual Hrs", "numeric"),
    MasterdataFieldSpec("start_date", "Start Date", "date"),
    MasterdataFieldSpec("end_date", "End Date", "date"),
    MasterdataFieldSpec("apr_2026", "Apr 2026", "numeric", aliases=("Apr-2026", "Apr-26")),
    MasterdataFieldSpec("may_2026", "May 2026", "numeric", aliases=("May-2026", "May-26")),
    MasterdataFieldSpec("jun_2026", "Jun 2026", "numeric", aliases=("Jun-2026", "Jun-26")),
    MasterdataFieldSpec("jul_2026", "Jul 2026", "numeric", aliases=("Jul-2026", "Jul-26")),
    MasterdataFieldSpec("aug_2026", "Aug 2026", "numeric", aliases=("Aug-2026", "Aug-26")),
    MasterdataFieldSpec("sep_2026", "Sep 2026", "numeric", aliases=("Sep-2026", "Sep-26")),
    MasterdataFieldSpec("oct_2026", "Oct 2026", "numeric", aliases=("Oct-2026", "Oct-26")),
    MasterdataFieldSpec("nov_2026", "Nov 2026", "numeric", aliases=("Nov-2026", "Nov-26")),
    MasterdataFieldSpec("dec_2026", "Dec 2026", "numeric", aliases=("Dec-2026", "Dec-26")),
    MasterdataFieldSpec("jan_2027", "Jan 2027", "numeric", aliases=("Jan-2027", "Jan-27")),
    MasterdataFieldSpec("feb_2027", "Feb 2027", "numeric", aliases=("Feb-2027", "Feb-27")),
    MasterdataFieldSpec("mar_2027", "Mar 2027", "numeric", aliases=("Mar-2027", "Mar-27")),
    MasterdataFieldSpec("fy", "FY", "numeric", aliases=("FY 2027", "FY2027")),
    MasterdataFieldSpec("project_name", "Project Name", "text", aliases=("Project",)),
    MasterdataFieldSpec("client_name", "Client Name", "text"),
    MasterdataFieldSpec(
        "ocn_number",
        "OCN Number",
        "text",
        aliases=("OCN", "OCN No", "OCN_NUMBER", "OCN Number"),
    ),
    MasterdataFieldSpec("practice_head", "Practice Head", "text"),
    MasterdataFieldSpec("bdm", "BDM", "text"),
    MasterdataFieldSpec("geo_head", "Geo Head", "text", aliases=("GeoHead", "Geo head")),
    MasterdataFieldSpec("vertical", "Vertical", "text"),
    MasterdataFieldSpec("horizontal", "Horizontal", "text"),
    MasterdataFieldSpec("q1", "Q1", "numeric"),
    MasterdataFieldSpec("q2", "Q2", "numeric"),
    MasterdataFieldSpec("q3", "Q3", "numeric"),
    MasterdataFieldSpec("q4", "Q4", "numeric"),
)

MASTERDATA_FIELD_BY_KEY = {field.key: field for field in MASTERDATA_FIELD_SPECS}
MASTERDATA_COLUMN_KEYS = tuple(field.key for field in MASTERDATA_FIELD_SPECS)
MASTERDATA_JSON_LABELS = {
    field.key: field.label for field in MASTERDATA_FIELD_SPECS
}

MASTERDATA_REQUIRED_COLUMNS: dict[DatasetType, tuple[str, ...]] = {
    "budget": ("customer_name", "ms_ps", "project_name"),
    "global_revenue": ("customer_name", "ms_ps", "resource_name", "project_name"),
    "forecast": ("ms_ps", "resource_name", "resource_id"),
}

MASTERDATA_REQUIRED_HEADER_THRESHOLD = 8
MASTERDATA_ROW_IDENTIFIER_FIELDS = (
    "resource_id",
    "ocn_number",
    "project_name",
    "resource_name",
    "customer_name",
)
MASTERDATA_DATE_FORMATS = (
    "%Y-%m-%d",
    "%d-%m-%Y",
    "%d/%m/%Y",
    "%m/%d/%Y",
    "%d-%b-%Y",
    "%d %b %Y",
    "%d.%m.%Y",
)
MASTERDATA_HEADER_DATE_FORMATS = (
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d",
    "%d-%b-%Y",
    "%d-%b-%y",
    "%b-%Y",
    "%b-%y",
    "%b %Y",
    "%B %Y",
    "%d/%m/%Y",
    "%m/%d/%Y",
)
MASTERDATA_MONTH_FIELD_BY_YEAR_MONTH: dict[tuple[int, int], str] = {
    (2026, 4): "apr_2026",
    (2026, 5): "may_2026",
    (2026, 6): "jun_2026",
    (2026, 7): "jul_2026",
    (2026, 8): "aug_2026",
    (2026, 9): "sep_2026",
    (2026, 10): "oct_2026",
    (2026, 11): "nov_2026",
    (2026, 12): "dec_2026",
    (2027, 1): "jan_2027",
    (2027, 2): "feb_2027",
    (2027, 3): "mar_2027",
}


def normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.strip().lower())


_MASTERDATA_HEADER_LOOKUP: dict[str, str] = {}
for field in MASTERDATA_FIELD_SPECS:
    for candidate in {field.label, field.key, *field.aliases}:
        normalized = normalize_header(candidate)
        if normalized:
            _MASTERDATA_HEADER_LOOKUP[normalized] = field.key


def parse_masterdata_workbook(
    path: Path,
    dataset_type: DatasetType,
) -> MasterdataParsedWorkbook:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        workbook = _parse_openxml_workbook(path)
    elif suffix == ".xls":
        workbook = _parse_legacy_xls_workbook(path)
    elif suffix == ".csv":
        workbook = _parse_csv_workbook(path)
    else:
        raise ValueError("Upload a workbook in .xlsx, .xls, or .csv format.")

    required_keys = set(MASTERDATA_REQUIRED_COLUMNS.get(dataset_type, ()))
    matched_keys = {
        _MASTERDATA_HEADER_LOOKUP[normalize_header(column)]
        for column in workbook.matched_columns
        if normalize_header(column) in _MASTERDATA_HEADER_LOOKUP
    }
    if "updated_customer" in matched_keys:
        matched_keys.add("customer_name")
    for row in workbook.rows:
        if not str(row.values.get("customer_name") or "").strip() and row.values.get("updated_customer"):
            row.values["customer_name"] = row.values["updated_customer"]
    missing_required_keys = sorted(required_keys - matched_keys)
    missing_required_columns = [
        MASTERDATA_JSON_LABELS[key] for key in missing_required_keys if key in MASTERDATA_JSON_LABELS
    ]

    return MasterdataParsedWorkbook(
        rows=workbook.rows,
        parsed_sheets=workbook.parsed_sheets,
        matched_columns=workbook.matched_columns,
        missing_required_columns=missing_required_columns,
        invalid_rows=[],
    )


def validate_masterdata_row(dataset_type: DatasetType, values: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    for key in MASTERDATA_REQUIRED_COLUMNS.get(dataset_type, ()):
        if values.get(key) in (None, ""):
            errors.append(f"Missing required value: {MASTERDATA_JSON_LABELS.get(key, key)}")

    return errors


def build_masterdata_business_key(dataset_type: DatasetType, values: dict[str, Any]) -> str:
    ms_ps = str(values.get("ms_ps") or "").strip().lower()
    resource_id = str(values.get("resource_id") or "").strip().lower()
    ocn = str(values.get("ocn_number") or "").strip().lower()
    project_name = str(values.get("project_name") or "").strip().lower()
    resource_name = str(values.get("resource_name") or "").strip().lower()
    customer_name = str(values.get("customer_name") or "").strip().lower()

    if dataset_type == "budget":
        if ms_ps == "ms" and ocn:
            return f"ms:{ocn}"
        if ms_ps == "ps" and resource_id:
            return f"ps:{resource_id}"
        if resource_id or ocn:
            return f"{ms_ps or 'na'}:{resource_id or ocn}"
        return f"budget:{project_name}:{customer_name or resource_name}"

    if dataset_type == "global_revenue":
        stable = resource_id or ocn or f"{project_name}:{customer_name or resource_name}"
        return f"global:{ms_ps or 'na'}:{stable}"

    stable = resource_id or ocn or f"{project_name}:{resource_name}"
    return f"forecast:{ms_ps or 'na'}:{stable}"


def _parse_openxml_workbook(path: Path) -> MasterdataParsedWorkbook:
    workbook = load_workbook(filename=path, data_only=True, read_only=True)
    rows: list[MasterdataParsedRow] = []
    parsed_sheets: list[str] = []
    matched_columns: set[str] = set()

    try:
        for worksheet in workbook.worksheets:
            sheet_rows, sheet_matches = _extract_sheet_rows_from_iterable(
                worksheet.iter_rows(values_only=True),
                worksheet.title,
            )
            if not sheet_rows:
                continue
            rows.extend(sheet_rows)
            parsed_sheets.append(worksheet.title)
            matched_columns.update(sheet_matches)
    finally:
        workbook.close()

    return MasterdataParsedWorkbook(
        rows=rows,
        parsed_sheets=parsed_sheets,
        matched_columns=sorted(matched_columns),
        missing_required_columns=[],
        invalid_rows=[],
    )


def _extract_sheet_rows_from_iterable(
    row_iterable: Any,
    sheet_name: str,
) -> tuple[list[MasterdataParsedRow], set[str]]:
    preview_rows: list[list[Any]] = []
    output: list[MasterdataParsedRow] = []
    matched_labels: set[str] = set()
    headers: list[Any] | None = None
    recognized_columns: dict[int, MasterdataFieldSpec] | None = None

    for row_number, raw_row_tuple in enumerate(row_iterable, start=1):
        raw_row = list(raw_row_tuple)

        if headers is None or recognized_columns is None:
            if len(preview_rows) < 25:
                preview_rows.append(raw_row)
            header_index, candidate_columns = _detect_header_row(preview_rows)
            has_candidate_header = (
                header_index is not None
                and len(candidate_columns) >= MASTERDATA_REQUIRED_HEADER_THRESHOLD
            )

            should_finalize_header = (
                has_candidate_header
                and header_index is not None
                and len(preview_rows) > header_index + 1
            ) or len(preview_rows) >= 25

            if not should_finalize_header:
                continue

            if not has_candidate_header or header_index is None:
                return [], set()

            headers = preview_rows[header_index]
            recognized_columns = candidate_columns
            matched_labels = {field.label for field in recognized_columns.values()}

            for preview_offset, preview_row in enumerate(preview_rows[header_index + 1 :]):
                preview_row_number = header_index + 2 + preview_offset
                parsed_row = _parse_sheet_row(
                    headers=headers,
                    recognized_columns=recognized_columns,
                    raw_row=preview_row,
                    sheet_name=sheet_name,
                    row_number=preview_row_number,
                )
                if parsed_row is not None:
                    output.append(parsed_row)
            continue

        parsed_row = _parse_sheet_row(
            headers=headers,
            recognized_columns=recognized_columns,
            raw_row=raw_row,
            sheet_name=sheet_name,
            row_number=row_number,
        )
        if parsed_row is not None:
            output.append(parsed_row)

    if headers is None or recognized_columns is None:
        header_index, candidate_columns = _detect_header_row(preview_rows)
        if (
            header_index is None
            or len(candidate_columns) < MASTERDATA_REQUIRED_HEADER_THRESHOLD
        ):
            return [], set()
        headers = preview_rows[header_index]
        recognized_columns = candidate_columns
        matched_labels = {field.label for field in recognized_columns.values()}

        for preview_offset, preview_row in enumerate(preview_rows[header_index + 1 :]):
            preview_row_number = header_index + 2 + preview_offset
            parsed_row = _parse_sheet_row(
                headers=headers,
                recognized_columns=recognized_columns,
                raw_row=preview_row,
                sheet_name=sheet_name,
                row_number=preview_row_number,
            )
            if parsed_row is not None:
                output.append(parsed_row)

    return output, matched_labels


def _parse_legacy_xls_workbook(path: Path) -> MasterdataParsedWorkbook:
    if xlrd is None:
        raise ValueError("The xlrd dependency is required for .xls workbook uploads.")

    workbook = xlrd.open_workbook(str(path), on_demand=True)
    rows: list[MasterdataParsedRow] = []
    parsed_sheets: list[str] = []
    matched_columns: set[str] = set()

    for sheet in workbook.sheets():
        raw_rows: list[list[Any]] = []
        for row_index in range(sheet.nrows):
            row_values: list[Any] = []
            for column_index in range(sheet.ncols):
                row_values.append(
                    _coerce_xls_cell_value(
                        workbook,
                        sheet.cell(row_index, column_index),
                    )
                )
            raw_rows.append(row_values)

        sheet_rows, sheet_matches = _extract_sheet_rows(raw_rows, sheet.name)
        if not sheet_rows:
            continue
        rows.extend(sheet_rows)
        parsed_sheets.append(sheet.name)
        matched_columns.update(sheet_matches)

    return MasterdataParsedWorkbook(
        rows=rows,
        parsed_sheets=parsed_sheets,
        matched_columns=sorted(matched_columns),
        missing_required_columns=[],
        invalid_rows=[],
    )


def _parse_csv_workbook(path: Path) -> MasterdataParsedWorkbook:
    rows: list[list[Any]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(handle, dialect)
        for row in reader:
            rows.append(list(row))

    sheet_rows, matched_columns = _extract_sheet_rows(rows, "CSV")
    return MasterdataParsedWorkbook(
        rows=sheet_rows,
        parsed_sheets=["CSV"] if sheet_rows else [],
        matched_columns=sorted(matched_columns),
        missing_required_columns=[],
        invalid_rows=[],
    )


def _coerce_xls_cell_value(workbook: Any, cell: Any) -> Any:
    if xlrd is None:
        return cell.value

    if cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate_as_datetime(cell.value, workbook.datemode)
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        numeric_value = float(cell.value)
        return int(numeric_value) if numeric_value.is_integer() else numeric_value
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK, xlrd.XL_CELL_ERROR}:
        return None
    return cell.value


def _extract_sheet_rows(
    raw_rows: list[list[Any]],
    sheet_name: str,
) -> tuple[list[MasterdataParsedRow], set[str]]:
    header_index, recognized_columns = _detect_header_row(raw_rows)
    if header_index is None or len(recognized_columns) < MASTERDATA_REQUIRED_HEADER_THRESHOLD:
        return [], set()

    headers = raw_rows[header_index]
    output: list[MasterdataParsedRow] = []
    matched_labels = {field.label for field in recognized_columns.values()}

    for row_offset, raw_row in enumerate(raw_rows[header_index + 1 :], start=header_index + 2):
        if _is_blank_row(raw_row):
            continue

        values: dict[str, Any] = {}
        raw_payload: dict[str, Any] = {}

        for column_index, header_value in enumerate(headers):
            if column_index >= len(raw_row):
                continue

            header_text = str(header_value or "").strip()
            if not header_text:
                continue

            raw_cell_value = raw_row[column_index]
            raw_payload[header_text] = _serialize_value(raw_cell_value)

            field = recognized_columns.get(column_index)
            if field is None:
                continue

            coerced_value = _coerce_field_value(field.kind, raw_cell_value)
            if coerced_value is not None:
                values[field.key] = coerced_value

        if not _is_meaningful_row(values):
            continue

        output.append(
            MasterdataParsedRow(
                sheet_name=sheet_name,
                row_number=row_offset,
                values=values,
                raw_payload=raw_payload,
            )
        )

    return output, matched_labels


def _parse_sheet_row(
    *,
    headers: list[Any],
    recognized_columns: dict[int, MasterdataFieldSpec],
    raw_row: list[Any],
    sheet_name: str,
    row_number: int,
) -> MasterdataParsedRow | None:
    if _is_blank_row(raw_row):
        return None

    values: dict[str, Any] = {}
    raw_payload: dict[str, Any] = {}

    for column_index, header_value in enumerate(headers):
        if column_index >= len(raw_row):
            continue

        header_text = str(header_value or "").strip()
        if not header_text:
            continue

        raw_cell_value = raw_row[column_index]
        raw_payload[header_text] = _serialize_value(raw_cell_value)

        field = recognized_columns.get(column_index)
        if field is None:
            continue

        coerced_value = _coerce_field_value(field.kind, raw_cell_value)
        if coerced_value is not None:
            values[field.key] = coerced_value

    if not _is_meaningful_row(values):
        return None

    return MasterdataParsedRow(
        sheet_name=sheet_name,
        row_number=row_number,
        values=values,
        raw_payload=raw_payload,
    )


def _detect_header_row(raw_rows: list[list[Any]]) -> tuple[int | None, dict[int, MasterdataFieldSpec]]:
    best_index: int | None = None
    best_columns: dict[int, MasterdataFieldSpec] = {}

    for row_index, row in enumerate(raw_rows[:25]):
        recognized_columns: dict[int, MasterdataFieldSpec] = {}
        for column_index, cell_value in enumerate(row):
            field_key = _resolve_header_field_key(cell_value)
            if not field_key:
                continue

            if any(existing.key == field_key for existing in recognized_columns.values()):
                continue

            recognized_columns[column_index] = MASTERDATA_FIELD_BY_KEY[field_key]

        if len(recognized_columns) > len(best_columns):
            best_index = row_index
            best_columns = recognized_columns

    return best_index, best_columns


def _resolve_header_field_key(value: Any) -> str | None:
    header_date = _parse_header_as_date(value)
    if header_date is not None:
        month_field = MASTERDATA_MONTH_FIELD_BY_YEAR_MONTH.get(
            (header_date.year, header_date.month),
        )
        if month_field:
            return month_field

    normalized = normalize_header(str(value or ""))
    if normalized:
        return _MASTERDATA_HEADER_LOOKUP.get(normalized)
    return None


def _parse_header_as_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    try:
        parsed = datetime.fromisoformat(text)
        return parsed.date()
    except ValueError:
        pass

    for pattern in MASTERDATA_HEADER_DATE_FORMATS:
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


def _is_blank_row(row: list[Any]) -> bool:
    return all(_is_empty_value(value) for value in row)


def _is_empty_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _coerce_field_value(kind: FieldKind, value: Any) -> Any:
    if _is_empty_value(value):
        return None
    if kind == "text":
        return _coerce_text_value(value)
    if kind == "numeric":
        return _coerce_numeric_value(value)
    return _coerce_date_value(value)


def _coerce_text_value(value: Any) -> str | None:
    if _is_empty_value(value):
        return None
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    text = str(value).strip()
    return text or None


def _coerce_numeric_value(value: Any) -> Decimal | None:
    if _is_empty_value(value):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))

    text = str(value).strip()
    if not text:
        return None

    negative = text.startswith("(") and text.endswith(")")
    cleaned = text.replace(",", "").replace("$", "").replace("%", "").strip()
    cleaned = cleaned.strip("()")
    cleaned = re.sub(r"[^0-9.\-]+", "", cleaned)
    if not cleaned:
        return None

    if negative and not cleaned.startswith("-"):
        cleaned = f"-{cleaned}"

    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def _coerce_date_value(value: Any) -> date | None:
    if _is_empty_value(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    for pattern in MASTERDATA_DATE_FORMATS:
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue

    return None


def _serialize_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def _is_meaningful_row(values: dict[str, Any]) -> bool:
    for value in values.values():
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return True
    return False
